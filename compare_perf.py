import openpyxl
import pandas as pd
from datetime import datetime
from pathlib import Path
import sys

def analyze_file(path_str):
    path = Path(path_str)
    if not path.exists():
        print(f"File not found: {path_str}")
        return None
    
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    header_row = 5
    agent_col_start = 0
    for r in range(1, 10):
        for c in range(1, 10):
            try:
                v = ws.cell(r, c).value
                if v and str(v).strip().lower() == 'user_name':
                    header_row = r
                    agent_col_start = c
                    break
            except: pass
        if agent_col_start: break
    
    if not agent_col_start:
        return None
        
    date_row = header_row - 1
    data_start_row = header_row + 1
    
    metric_cols = []
    for c in range(agent_col_start + 1, max_col + 1):
        v = ws.cell(header_row, c).value
        if v and str(v).strip().upper() in ('GADD', 'ADS'):
            metric_cols.append(c)
            
    col_info = {}
    last_date = None
    for c in metric_cols:
        d_val = ws.cell(date_row, c).value
        if d_val:
            if isinstance(d_val, datetime):
                last_date = d_val.date()
            else:
                try: 
                    last_date = pd.to_datetime(str(d_val)).date()
                except: pass
        
        m_type = str(ws.cell(header_row, c).value).strip().upper()
        if last_date:
            col_info[c] = {'date': last_date, 'type': m_type}
        
    results = []
    for r in range(data_start_row, max_row + 1):
        uname = ws.cell(r, agent_col_start).value
        if not uname or str(uname).strip() == "" or str(uname).strip() == "0":
            continue
            
        for c, info in col_info.items():
            val = ws.cell(r, c).value or 0
            try: val = float(val)
            except: val = 0
            results.append({'date': info['date'], 'type': info['type'], 'value': val})
            
    df = pd.DataFrame(results)
    if df.empty: return pd.DataFrame()
    summary = df.groupby(['date', 'type'])['value'].sum().unstack(fill_value=0)
    return summary

f1 = 'inputs/GLOBALE PERFORMANCE 20260413.xlsx'
f2 = 'inputs/GLOBALE PERFORMANCE 20260419.xlsx'

s1 = analyze_file(f1)
s2 = analyze_file(f2)

if s1 is not None:
    print(f"Summary for {f1}:")
    print(s1.tail(10))
else:
    print(f"Could not parse {f1}")

if s2 is not None:
    print(f"\nSummary for {f2}:")
    print(s2.tail(10))
else:
    print(f"Could not parse {f2}")

import pandas as pd
target_dates = [pd.to_datetime('2026-04-11').date(), pd.to_datetime('2026-04-12').date()]
print("\nValidation of hypothesis:")
for d in target_dates:
    print(f"Date: {d}")
    v1_gadd = s1.loc[d, 'GADD'] if s1 is not None and d in s1.index and 'GADD' in s1.columns else 0
    v1_ads = s1.loc[d, 'ADS'] if s1 is not None and d in s1.index and 'ADS' in s1.columns else 0
    v2_gadd = s2.loc[d, 'GADD'] if s2 is not None and d in s2.index and 'GADD' in s2.columns else 0
    v2_ads = s2.loc[d, 'ADS'] if s2 is not None and d in s2.index and 'ADS' in s2.columns else 0
    
    print(f"  File 20260413: GADD={v1_gadd}, ADS={v1_ads}")
    print(f"  File 20260419: GADD={v2_gadd}, ADS={v2_ads}")
    
    if v1_gadd == 0 and v1_ads == 0 and (v2_gadd != 0 or v2_ads != 0):
        print(f"  ==> Hypothesis CONFIRMED: Metrics are zero in 0413 but non-zero in 0419.")
    else:
        print(f"  ==> Hypothesis NOT confirmed.")