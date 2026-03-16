import openpyxl

file_path = r"d:\LKA\Perf_commissions\Point Financier Superviseur BA Dec 2025.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)

# Afficher les feuilles
print("--- Onglets disponibles ---")
print(wb.sheetnames)

# Cibler le Recap
recap_sheet = next((s for s in wb.sheetnames if 'recap' in s.lower()), wb.sheetnames[0])
ws = wb[recap_sheet]

print(f"\n--- Analyse de l'onglet: {recap_sheet} ---")
print(f"Dimensions: {ws.dimensions} (Lignes: {ws.max_row}, Colonnes: {ws.max_column})")

print("\n--- En-têtes (Ligne 1 à 3 approx) ---")
for r in range(1, min(4, ws.max_row + 1)):
    row_vals = [str(ws.cell(r, c).value) for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None]
    if row_vals:
        print(f"Ligne {r}:", " | ".join(row_vals))

print("\n--- Formules clés détectées (Échantillon) ---")
formula_count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.data_type == 'f':  # Si c'est une formule
            print(f"Cellule {cell.coordinate} : {cell.value}")
            formula_count += 1
            if formula_count >= 15:
                break
    if formula_count >= 15:
        break