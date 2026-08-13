"""
scripts/05_commission.py -- Génération du fichier Excel de commissions
Architecture "Data-Driven"
Design UX Mobile-first avec en-tête figé, matrice tarifaire et colonnes épurées !
"""

import sys
import argparse
import json
from collections import Counter
from pathlib import Path
from datetime import datetime, date, timedelta

import pandas as pd
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT = Path(__file__).parent.parent
OUTPUTS = ROOT / "outputs"
LOGS = ROOT / "logs"
COMMISSION_STATE_FILE = LOGS / "commission_retry_state.json"

sys.path.insert(0, str(ROOT))
from connections.config import (
    MYSQL_DATABASE, TABLE_DAILY_GADD, TABLE_DAILY_ADS,
    AUTO_MODE, AUTO_DAYS_RANGE, MANUAL_END_DATE, MANUAL_RANGE_DAYS,
    TOTAL_WARNINGS_FILE,
)
from connections.connect import make_engine

# ─── CONSTANTES DE STYLE ───────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUB_HEADER_FONT = Font(bold=True, color="000000", size=10)
SUB_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CURRENCY_FMT = '#,##0'
IDENTITY_COLUMNS = ['USER NAME', 'Superviseur', 'AGENT_NAME', 'MSISDN', 'CANAL']

# ─── NORMALISATION DES CANAUX ───────────────────────────────────────────────
# Argent sensible : les noms de canaux (real_channel) arrivent avec des variantes
# (espaces, underscores, pluriels, casse). On centralise ici la correspondance
# vers le canal CANONIQUE (= type_agent dans commission_tarifs).
def normalize_channel_key(channel) -> str:
    if channel is None or (isinstance(channel, float) and pd.isna(channel)):
        return ""
    return " ".join(str(channel).strip().upper().split())

CHANNEL_CANONICAL = {
    "BA": "BA",
    "BA CLASSIQUE": "BA",
    "BA CLASSIQUES": "BA",
    "BA AGENCE": "BA_AGENCE",
    "BA_AGENCE": "BA_AGENCE",
    "ANIMATION PICK-UP": "Animation Pick-up",
    "ANIMATION POS": "Animation POS",
    "MA": "MA",
}
# Canaux explicitement NON commissionnes : exclus du fichier sans generer d'alerte.
EXCLUDED_CHANNELS = {"KIOSQUES", "KIOSQUE", "POS"}

def canonical_channel(channel):
    """Renvoie le canal canonique paye, ou None si inconnu/non paye."""
    return CHANNEL_CANONICAL.get(normalize_channel_key(channel))

# ─── HELPERS ───────────────────────────────────────────────────────────────────
def format_date_fr(d: date) -> str:
    jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    mois = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    return f"{jours[d.weekday()]} {d.day} {mois[d.month-1]} {d.year}"

def format_day_header_fr(d: date) -> str:
    jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
    mois = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    return f"{jours[d.weekday()]} {d.day} {mois[d.month-1]}"

def iter_dates(date_start: date, date_end: date):
    current = date_start
    while current <= date_end:
        yield current
        current += timedelta(days=1)

def load_commission_state():
    if not COMMISSION_STATE_FILE.exists():
        return {"pending_retries": []}

    try:
        return json.loads(COMMISSION_STATE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {"pending_retries": []}


def parse_state_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None

def save_commission_state(state):
    LOGS.mkdir(exist_ok=True)

    existing_state = load_commission_state()

    seen = set()
    pending_retries = []
    raw_pending_retries = state.get(
        "pending_retries",
        existing_state.get("pending_retries", []),
    )
    for item in raw_pending_retries:
        retry_date = item.get("date")
        metric = item.get("metric")
        if not retry_date or not metric:
            continue
        key = (retry_date, metric)
        if key in seen:
            continue
        seen.add(key)
        pending_retries.append({
            "date": retry_date,
            "metric": metric,
            "reason": item.get("reason", ""),
        })

    payload = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "pending_retries": sorted(
            pending_retries,
            key=lambda item: (item["date"], item["metric"]),
        ),
    }

    anchor = state.get("daily_recompute_anchor", existing_state.get("daily_recompute_anchor"))
    if isinstance(anchor, dict):
        processing_day = anchor.get("processing_day")
        start_date = parse_state_date(anchor.get("start_date"))
        end_date = parse_state_date(anchor.get("end_date"))
        if processing_day and start_date:
            payload["daily_recompute_anchor"] = {
                "processing_day": processing_day,
                "start_date": start_date.isoformat(),
            }
            if end_date:
                payload["daily_recompute_anchor"]["end_date"] = end_date.isoformat()

    last_mail_context = state.get("last_mail_context", existing_state.get("last_mail_context"))
    if isinstance(last_mail_context, dict):
        source_start_date = parse_state_date(last_mail_context.get("source_start_date"))
        source_end_date = parse_state_date(last_mail_context.get("source_end_date"))
        commission_start_date = parse_state_date(last_mail_context.get("commission_start_date"))
        commission_end_date = parse_state_date(last_mail_context.get("commission_end_date"))
        if source_start_date and source_end_date and commission_start_date and commission_end_date:
            payload["last_mail_context"] = {
                "source_start_date": source_start_date.isoformat(),
                "source_end_date": source_end_date.isoformat(),
                "commission_start_date": commission_start_date.isoformat(),
                "commission_end_date": commission_end_date.isoformat(),
            }

    COMMISSION_STATE_FILE.write_text(
        json.dumps(payload, indent=2, ensure_ascii=True),
        encoding="utf-8",
    )


def load_daily_recompute_anchor(processing_day: date | None = None):
    processing_day = processing_day or date.today()
    anchor = load_commission_state().get("daily_recompute_anchor", {})
    if not isinstance(anchor, dict):
        return None
    if anchor.get("processing_day") != processing_day.isoformat():
        return None
    return parse_state_date(anchor.get("start_date"))


def update_daily_recompute_anchor(start_date: date, end_date: date, processing_day: date | None = None):
    processing_day = processing_day or date.today()
    existing_start = load_daily_recompute_anchor(processing_day)
    if existing_start and existing_start < start_date:
        start_date = existing_start

    save_commission_state({
        "daily_recompute_anchor": {
            "processing_day": processing_day.isoformat(),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
        }
    })
    return start_date


def load_last_mail_context():
    raw_context = load_commission_state().get("last_mail_context", {})
    if not isinstance(raw_context, dict):
        return None

    source_start_date = parse_state_date(raw_context.get("source_start_date"))
    source_end_date = parse_state_date(raw_context.get("source_end_date"))
    commission_start_date = parse_state_date(raw_context.get("commission_start_date"))
    commission_end_date = parse_state_date(raw_context.get("commission_end_date"))

    if not all([source_start_date, source_end_date, commission_start_date, commission_end_date]):
        return None

    return {
        "source_start_date": source_start_date,
        "source_end_date": source_end_date,
        "commission_start_date": commission_start_date,
        "commission_end_date": commission_end_date,
    }


def get_processed_source_range(processing_day: date | None = None):
    processing_day = processing_day or date.today()
    suffix = processing_day.strftime("%Y-%m-%d")
    candidate_files = [
        OUTPUTS / f"gadd_long_{suffix}.xlsx",
        OUTPUTS / f"ads_long_{suffix}.xlsx",
    ]

    all_dates = []
    for path in candidate_files:
        if not path.exists():
            continue
        try:
            df = pd.read_excel(path, usecols=["perf_date"])
        except Exception:
            continue
        if "perf_date" not in df.columns:
            continue
        parsed_dates = pd.to_datetime(df["perf_date"], errors="coerce").dropna()
        all_dates.extend(parsed_dates.dt.date.tolist())

    if not all_dates:
        return None

    return {
        "source_start_date": min(all_dates),
        "source_end_date": max(all_dates),
    }


def save_last_mail_context(commission_start_date: date, commission_end_date: date, processing_day: date | None = None):
    current_source_range = get_processed_source_range(processing_day)
    if not current_source_range:
        return None

    save_commission_state({
        "last_mail_context": {
            "source_start_date": current_source_range["source_start_date"].isoformat(),
            "source_end_date": current_source_range["source_end_date"].isoformat(),
            "commission_start_date": commission_start_date.isoformat(),
            "commission_end_date": commission_end_date.isoformat(),
        }
    })
    return current_source_range

def find_last_commission_date():
    import os
    import re

    pattern_old = r"commission_.*?(\d{4}-\d{2}-\d{2})\.xlsx$"
    pattern_new = r"Variables .* (\d{2}-\d{2}-\d{4})\.xlsx$"
    pattern_unified = r"Commission LKA .* (\d{2}-\d{2}-\d{4})\.xlsx$"

    dates_found = []
    if os.path.exists(OUTPUTS):
        for filename in os.listdir(OUTPUTS):
            match_old = re.search(pattern_old, filename)
            if match_old:
                dates_found.append(datetime.strptime(match_old.group(1), "%Y-%m-%d").date())

            match_new = re.search(pattern_new, filename)
            if match_new:
                dates_found.append(datetime.strptime(match_new.group(1), "%d-%m-%Y").date())

            match_unified = re.search(pattern_unified, filename)
            if match_unified:
                dates_found.append(datetime.strptime(match_unified.group(1), "%d-%m-%Y").date())

    return max(dates_found) if dates_found else None

def get_latest_source_date(engine):
    sql = text(f"""
        SELECT GREATEST(
            COALESCE((SELECT MAX(perf_date) FROM {TABLE_DAILY_GADD}), '1900-01-01'),
            COALESCE((SELECT MAX(perf_date) FROM {TABLE_DAILY_ADS}), '1900-01-01')
        )
    """)
    with engine.connect() as conn:
        row = conn.execute(sql).fetchone()
    if not row or not row[0]:
        return None
    return pd.to_datetime(row[0]).date()

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 25)

def fetch_data(engine, view_name, val_col, comm_col, date_start, date_end):
    sql = text(f"""
        SELECT 
            user_name,
            superviseur AS nom_prenom_superviseur,
            agent_name,
            msisdn_momo,
            real_channel,
            perf_date,
            periode_nom,
            taux_{val_col}_applique AS pu,
            SUM({val_col}) AS nb_total,
            SUM({comm_col}) AS amt_total,
            SUM({comm_col}_t1) AS amt_t1,
            SUM({comm_col}_t2) AS amt_t2
        FROM {view_name}
        WHERE perf_date BETWEEN :ds AND :de
        GROUP BY 
            user_name, superviseur, agent_name, msisdn_momo, real_channel, perf_date, periode_nom, taux_{val_col}_applique
    """)
    with engine.connect() as conn:
        df = pd.read_sql(sql, conn, params={"ds": date_start, "de": date_end})
    if not df.empty:
        df['perf_date'] = pd.to_datetime(df['perf_date']).dt.date
    return df

def get_daily_metric_totals(engine, date_start, date_end):
    totals = {
        current_date: {"GADD": 0, "ADS": 0}
        for current_date in iter_dates(date_start, date_end)
    }

    sql_gadd = text(f"""
        SELECT perf_date, SUM(gadd) AS total_qty
        FROM {TABLE_DAILY_GADD}
        WHERE perf_date BETWEEN :ds AND :de
        GROUP BY perf_date
    """)
    sql_ads = text(f"""
        SELECT perf_date, SUM(ads) AS total_qty
        FROM {TABLE_DAILY_ADS}
        WHERE perf_date BETWEEN :ds AND :de
        GROUP BY perf_date
    """)

    with engine.connect() as conn:
        gadd_rows = conn.execute(sql_gadd, {"ds": date_start, "de": date_end}).fetchall()
        ads_rows = conn.execute(sql_ads, {"ds": date_start, "de": date_end}).fetchall()

    for perf_date, total_qty in gadd_rows:
        totals[pd.to_datetime(perf_date).date()]["GADD"] = int(total_qty or 0)
    for perf_date, total_qty in ads_rows:
        totals[pd.to_datetime(perf_date).date()]["ADS"] = int(total_qty or 0)

    return totals


def verify_totals(engine, df_gadd_raw, df_ads_raw, df_all_gadd, df_all_ads,
                  day_dates, date_start, date_end, all_group_channels):
    """Compare les totaux du fichier Excel (pivot) avec ceux de la base.
    Écrit un fichier JSON de warnings et retourne la liste."""
    warnings = []
    info_messages = []

    # 1. Totaux depuis le pivot (ce qui sera dans le fichier Excel)
    pivot_gadd = {}
    pivot_ads = {}
    for day_date in day_dates:
        day_label = format_day_header_fr(day_date)
        if not df_all_gadd.empty and day_label in df_all_gadd.columns:
            pivot_gadd[day_date] = int(df_all_gadd[day_label].sum())
        if not df_all_ads.empty and day_label in df_all_ads.columns:
            pivot_ads[day_date] = int(df_all_ads[day_label].sum())

    # 2. Totaux depuis les vues brutes (filtrées par canaux du fichier)
    raw_gadd = {}
    raw_ads = {}
    def _paid_mask(df_raw):
        if 'canal_norm' in df_raw.columns:
            return df_raw['canal_norm'].isin(all_group_channels)
        return df_raw['real_channel'].map(canonical_channel).isin(all_group_channels)
    if not df_gadd_raw.empty and 'real_channel' in df_gadd_raw.columns:
        df_g_filt = df_gadd_raw[_paid_mask(df_gadd_raw)]
        raw_gadd = df_g_filt.groupby('perf_date')['nb_total'].sum().to_dict() if not df_g_filt.empty else {}
    if not df_ads_raw.empty and 'real_channel' in df_ads_raw.columns:
        df_a_filt = df_ads_raw[_paid_mask(df_ads_raw)]
        raw_ads = df_a_filt.groupby('perf_date')['nb_total'].sum().to_dict() if not df_a_filt.empty else {}

    # 3. Coherence interne : pivot vs vues filtrées (si écart = bug)
    for day_date in day_dates:
        pg = pivot_gadd.get(day_date, 0)
        pa = pivot_ads.get(day_date, 0)
        rg = raw_gadd.get(day_date, 0)
        ra = raw_ads.get(day_date, 0)

        if pg != rg:
            warnings.append({
                "date": day_date.isoformat(),
                "metric": "GADD",
                "file_total": pg,
                "db_total": rg,
                "diff": pg - rg,
                "message": f"GADD {format_date_fr(day_date)} : pivot={pg}, vues={rg}, écart={pg - rg:+d}"
            })
        if pa != ra:
            warnings.append({
                "date": day_date.isoformat(),
                "metric": "ADS",
                "file_total": pa,
                "db_total": ra,
                "diff": pa - ra,
                "message": f"ADS {format_date_fr(day_date)} : pivot={pa}, vues={ra}, écart={pa - ra:+d}"
            })

    # 4. Canaux exclus : vues filtrées vs base globale (info, pas alerte)
    db_totals = get_daily_metric_totals(engine, date_start, date_end)
    excluded_by_date = {}
    for day_date in day_dates:
        rg = raw_gadd.get(day_date, 0)
        ra = raw_ads.get(day_date, 0)
        dg = db_totals.get(day_date, {}).get("GADD", 0)
        da = db_totals.get(day_date, {}).get("ADS", 0)
        g_exc = dg - rg
        a_exc = da - ra
        if g_exc > 0 or a_exc > 0:
            excluded_by_date[day_date.isoformat()] = {"GADD": g_exc, "ADS": a_exc}

    if excluded_by_date:
        info_messages.append(f"Canaux exclus du fichier : {len(excluded_by_date)} jour(s) avec données hors FILE_GROUPS.")

    # 4bis. CONTROLE D'INTEGRITE DES CANAUX (argent sensible), par canal sur la base brute :
    #   - canal explicitement exclu (KIOSQUES/POS) : ignore (pas d'alerte)
    #   - canal paye connu avec des unites a TAUX 0 : non payees -> alerte
    #   - canal INCONNU avec du volume : possible faute d'ecriture d'un bon canal -> alerte
    for label, df_raw in (("GADD", df_gadd_raw), ("ADS", df_ads_raw)):
        if df_raw.empty or 'real_channel' not in df_raw.columns:
            continue
        agg = df_raw.groupby('real_channel', dropna=False).agg(
            qty=('nb_total', 'sum'), amount=('amt_total', 'sum')
        ).reset_index()
        for _, row in agg.iterrows():
            channel = row['real_channel']
            qty = int(row['qty'] or 0)
            amount = int(row['amount'] or 0)
            if qty <= 0:
                continue
            key = normalize_channel_key(channel)
            if key in EXCLUDED_CHANNELS:
                continue
            canon = CHANNEL_CANONICAL.get(key)
            if canon is None:
                sub = df_raw[df_raw['real_channel'].isna()] if pd.isna(channel) else df_raw[df_raw['real_channel'] == channel]
                users = sorted(set(str(u) for u in sub['user_name'].dropna().unique())) if 'user_name' in sub.columns else []
                users_str = ", ".join(users[:10]) + (f", ... (+{len(users) - 10} autre(s))" if len(users) > 10 else "") if users else "aucun user_name identifiable"
                warnings.append({
                    "metric": label, "channel": str(channel), "qty": qty, "amount": amount,
                    "message": (
                        f"{label} : canal INCONNU '{channel}' avec {qty} unite(s) (montant {amount}). "
                        f"Ni canal paye connu ni exclu -> verifier une faute d'ecriture (argent a risque). "
                        f"Agent(s) concerne(s) : {users_str}."
                    ),
                })
                continue
            sub = df_raw[df_raw['real_channel'] == channel]
            zero = sub[(sub['nb_total'] > 0) & (sub['pu'].fillna(0) == 0)]
            if not zero.empty:
                lost_qty = int(zero['nb_total'].sum())
                warnings.append({
                    "metric": label, "channel": str(channel), "qty": lost_qty, "amount": 0,
                    "message": (
                        f"{label} : canal '{channel}' (=> {canon}) a {lost_qty} unite(s) a TAUX 0 "
                        f"= NON payees. Tarif manquant pour ces dates/jours -> a verifier."
                    ),
                })

    # 4ter. Doublons de casse sur user_name (meme agent, ecritures differentes a la source).
    dup_names = {}
    for df_raw in (df_gadd_raw, df_ads_raw):
        if df_raw.empty or 'user_name' not in df_raw.columns:
            continue
        tmp = df_raw[['user_name']].dropna().copy()
        tmp['k'] = tmp['user_name'].astype(str).str.strip().str.lower()
        for k, names in tmp.groupby('k')['user_name']:
            variants = set(str(n).strip() for n in names)
            if len(variants) > 1:
                dup_names[k] = set(dup_names.get(k, set())) | variants
    for k, variants in sorted(dup_names.items()):
        warnings.append({
            "metric": "DOUBLON", "channel": "", "qty": len(variants), "amount": 0,
            "message": (
                f"Doublon de casse user_name : {' / '.join(sorted(variants))} = meme agent. "
                f"Fusionne dans le fichier, a corriger a la source (daily_gadd)."
            ),
        })

    # 5. Sauvegarde (warnings uniquement pour les écarts internes)
    Path(TOTAL_WARNINGS_FILE).parent.mkdir(parents=True, exist_ok=True)
    with open(TOTAL_WARNINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(warnings, f, ensure_ascii=False, indent=2)

    if warnings:
        print(f"\n⚠ {len(warnings)} anomalie(s) détectée(s) (totaux / canaux / doublons) :")
        for w in warnings:
            print(f"  {w['message']}")
    else:
        print("\n✅ Totaux, canaux et identifiants cohérents.")

    if info_messages:
        for m in info_messages:
            print(f"  ℹ {m}")

    if excluded_by_date:
        print(f"  ℹ Canaux exclus par jour :")
        for d, v in sorted(excluded_by_date.items()):
            parts = []
            if v["GADD"] > 0: parts.append(f"GADD +{v['GADD']}")
            if v["ADS"] > 0: parts.append(f"ADS +{v['ADS']}")
            print(f"     {d}: {', '.join(parts)}")

    return warnings

def has_any_activity(totals):
    return (totals.get("GADD", 0) + totals.get("ADS", 0)) > 0

def detect_pending_retries(daily_totals):
    pending = []
    sorted_totals = sorted(daily_totals.items())
    zero_run = []

    def flush_zero_run(next_totals=None, trailing=False):
        nonlocal zero_run
        if not zero_run:
            return

        first_idx = zero_run[0][0]
        prev_totals = sorted_totals[first_idx - 1][1] if first_idx > 0 else None
        actif_avant = bool(prev_totals) and has_any_activity(prev_totals)

        encadre = actif_avant and bool(next_totals) and has_any_activity(next_totals)
        # BORD DROIT. Une serie de zeros en FIN de plage n'a pas de jour suivant
        # PAR CONSTRUCTION -- exiger `next_totals` la rendait indetectable. Or
        # c'est le cas le plus dangereux : le dernier jour est justement celui
        # qu'on commissionne. Le 2026-08-11 est passe par ce trou (2177 agents
        # a 0, GADD et ADS, apres un 10/08 a 3981) et un classeur de
        # commissions vide est parti par mail, job vert.
        borde = trailing and actif_avant

        if encadre or borde:
            raison = ("GADD et ADS a 0 entre deux jours actifs" if encadre
                      else "GADD et ADS a 0 en fin de plage, apres un jour actif")
            for _, retry_date in zero_run:
                pending.append({
                    "date": retry_date.isoformat(),
                    "metric": "GADD+ADS",
                    "reason": raison,
                })

        zero_run = []

    for idx, (perf_date, totals) in enumerate(sorted_totals):
        gadd_total = totals.get("GADD", 0)
        ads_total = totals.get("ADS", 0)

        if gadd_total == 0 and ads_total == 0:
            zero_run.append((idx, perf_date))
            continue

        flush_zero_run(totals)

        if gadd_total > 0 and ads_total == 0:
            pending.append({
                "date": perf_date.isoformat(),
                "metric": "ADS",
                "reason": "ADS total a 0 alors que GADD est positif",
            })
        elif ads_total > 0 and gadd_total == 0:
            pending.append({
                "date": perf_date.isoformat(),
                "metric": "GADD",
                "reason": "GADD total a 0 alors que ADS est positif",
            })

    # La boucle ne vidait JAMAIS `zero_run` en sortant : une plage se terminant
    # par des zeros les jetait en silence. Second defaut du meme cas.
    flush_zero_run(trailing=True)

    return pending


def detect_retry_dates_in_range(engine, date_start, date_end):
    if date_start > date_end:
        return []
    return detect_pending_retries(get_daily_metric_totals(engine, date_start, date_end))

def refresh_commission_state(engine, date_start, date_end):
    state = load_commission_state()
    current_range_dates = {current_date.isoformat() for current_date in iter_dates(date_start, date_end)}
    remaining_pending = [
        item for item in state.get("pending_retries", [])
        if item.get("date") not in current_range_dates
    ]

    pending_for_current_range = detect_retry_dates_in_range(engine, date_start, date_end)
    remaining_pending.extend(pending_for_current_range)
    save_commission_state({"pending_retries": remaining_pending})
    return pending_for_current_range

def load_pending_retry_dates(engine, end_date):
    """Dates en attente dont la donnee manquante est ENFIN arrivee en base.

    Une date encore incomplete reste en attente SANS etirer la periode de
    calcul : on ne la re-inclut dans le range que le jour ou sa metrique
    manquante est reellement presente, pour la recalculer une seule fois."""
    resolved_dates = []
    for item in load_commission_state().get("pending_retries", []):
        try:
            retry_date = datetime.strptime(item["date"], "%Y-%m-%d").date()
        except Exception:
            continue
        if retry_date > end_date:
            continue
        metric = item.get("metric", "")
        totals = get_daily_metric_totals(engine, retry_date, retry_date).get(retry_date, {})
        gadd_total = totals.get("GADD", 0)
        ads_total = totals.get("ADS", 0)
        if metric == "ADS" and ads_total > 0:
            resolved_dates.append(retry_date)
        elif metric == "GADD" and gadd_total > 0:
            resolved_dates.append(retry_date)
        elif metric == "GADD+ADS" and (gadd_total > 0 or ads_total > 0):
            resolved_dates.append(retry_date)
    return sorted(set(resolved_dates))

def get_active_periods(engine, date_start, date_end, group_channels=None):
    where_clauses = ["date_debut <= :de", "date_fin >= :ds"]
    params = {"ds": date_start, "de": date_end}

    if group_channels:
        channel_params = []
        for idx, channel in enumerate(group_channels):
            key = f"channel_{idx}"
            channel_params.append(f":{key}")
            params[key] = channel
        where_clauses.append(f"type_agent IN ({', '.join(channel_params)})")

    sql = text(f"""
        SELECT
            periode_nom,
            MIN(COALESCE(jour_debut, 99)) AS jour_debut,
            MIN(COALESCE(jour_fin, 99)) AS jour_fin
        FROM commission_tarifs
        WHERE {' AND '.join(where_clauses)}
        GROUP BY periode_nom
        ORDER BY jour_debut, jour_fin, periode_nom
    """)

    with engine.connect() as conn:
        rows = conn.execute(sql, params).fetchall()

    ordered_rows = sorted(
        rows,
        key=lambda row: (
            8 if row[1] == 1 else row[1],
            8 if row[2] == 1 else row[2],
            row[0],
        ),
    )

    return [row[0] for row in ordered_rows if row[0]]

def get_date_period_map(engine, date_start, date_end, group_channels=None, frames=None):
    frames = frames or []
    date_period_map = {}

    for frame in frames:
        if frame is None or frame.empty or 'perf_date' not in frame.columns or 'periode_nom' not in frame.columns:
            continue

        temp = frame[['perf_date', 'periode_nom']].copy()
        temp['perf_date'] = pd.to_datetime(temp['perf_date']).dt.date
        temp['periode_nom'] = temp['periode_nom'].fillna('').astype(str).str.strip()
        temp = temp[temp['periode_nom'] != '']
        temp = temp[temp['periode_nom'] != 'Autre']

        if temp.empty:
            continue

        for perf_date, values in temp.groupby('perf_date')['periode_nom']:
            periods = sorted(set(values.tolist()))
            if periods:
                date_period_map[perf_date] = " / ".join(periods)

    if len(date_period_map) == len(list(iter_dates(date_start, date_end))):
        return date_period_map

    where_clauses = ["date_debut <= :de", "date_fin >= :ds"]
    params = {"ds": date_start, "de": date_end}

    if group_channels:
        channel_params = []
        for idx, channel in enumerate(group_channels):
            key = f"period_channel_{idx}"
            channel_params.append(f":{key}")
            params[key] = channel
        where_clauses.append(f"type_agent IN ({', '.join(channel_params)})")

    sql = text(f"""
        SELECT DISTINCT periode_nom, date_debut, date_fin, jour_debut, jour_fin
        FROM commission_tarifs
        WHERE {' AND '.join(where_clauses)}
    """)

    with engine.connect() as conn:
        rows = conn.execute(sql, params).fetchall()

    for current_date in iter_dates(date_start, date_end):
        if current_date in date_period_map:
            continue

        matched_periods = []
        mysql_day = ((current_date.weekday() + 1) % 7) + 1
        for periode_nom, date_debut, date_fin, jour_debut, jour_fin in rows:
            if not (date_debut <= current_date <= date_fin):
                continue
            if jour_debut is None or jour_fin is None:
                matched_periods.append(periode_nom)
                continue
            if jour_debut <= mysql_day <= jour_fin:
                matched_periods.append(periode_nom)

        matched_periods = sorted(set([period for period in matched_periods if period]))
        date_period_map[current_date] = " / ".join(matched_periods) if matched_periods else "Autre"

    return date_period_map

def build_period_blocks(day_dates, date_period_map):
    blocks = []
    current_period = None
    current_dates = []

    for day_date in day_dates:
        period_name = date_period_map.get(day_date, 'Autre')
        if current_period is None or period_name == current_period:
            current_period = period_name
            current_dates.append(day_date)
            continue

        blocks.append({"period": current_period, "dates": current_dates})
        current_period = period_name
        current_dates = [day_date]

    if current_dates:
        blocks.append({"period": current_period, "dates": current_dates})

    period_counts = Counter(block['period'] for block in blocks)
    seen_counts = {}
    for block in blocks:
        period_name = block['period']
        seen_counts[period_name] = seen_counts.get(period_name, 0) + 1
        suffix = "" if period_counts[period_name] == 1 else f" {seen_counts[period_name]}"
        block['total_label'] = f"TOTAL {period_name}{suffix}"
        block['amount_label'] = f"Mnt {period_name}{suffix}"
        # Decomposition des paliers (uniquement le bloc Dimanche) : <=10 et >10 adds.
        block['is_sunday'] = (period_name == 'Dimanche')
        if block['is_sunday']:
            block['amount_label_t1'] = f"Mnt {period_name}{suffix} <= 10"
            block['amount_label_t2'] = f"Mnt {period_name}{suffix} > 10"

    return blocks

def pivot_data(df, day_dates, period_blocks, qty_label="Add"):
    if df.empty:
        return pd.DataFrame()

    records = []
    # Dedup par agent insensible a la casse du user_name (la collation MySQL est deja
    # insensible a la casse : 'Octavie.Dadjo' et 'OCtavie.Dadjo' sont le MEME agent).
    df = df.copy()
    df['_user_key'] = df['user_name'].astype(str).str.strip().str.lower()
    grouped = df.groupby('_user_key', dropna=False)

    for _user_key, group in grouped:
        group = group.copy()
        group['perf_date'] = pd.to_datetime(group['perf_date']).dt.date
        rep = group.iloc[0]

        r = {
            'USER NAME': str(rep['user_name']).strip(),
            'Superviseur': rep['nom_prenom_superviseur'],
            'AGENT_NAME': rep['agent_name'],
            'MSISDN': rep['msisdn_momo'],
            'CANAL': rep['real_channel'],
            f'TOTAL {qty_label}': group['nb_total'].sum(),
            'TOTAL A PAYER': group['amt_total'].sum()
        }

        for day_date in day_dates:
            day_label = format_day_header_fr(day_date)
            day_group = group[group['perf_date'] == day_date]
            r[day_label] = day_group['nb_total'].sum() if not day_group.empty else 0

        for block in period_blocks:
            block_group = group[group['perf_date'].isin(block['dates'])]
            r[block['total_label']] = block_group['nb_total'].sum() if not block_group.empty else 0
            r[block['amount_label']] = block_group['amt_total'].sum() if not block_group.empty else 0
            if block.get('is_sunday'):
                t1 = block_group['amt_t1'].sum() if (not block_group.empty and 'amt_t1' in block_group.columns) else 0
                t2 = block_group['amt_t2'].sum() if (not block_group.empty and 'amt_t2' in block_group.columns) else 0
                r[block['amount_label_t1']] = t1
                r[block['amount_label_t2']] = t2

        records.append(r)
        
    res_df = pd.DataFrame(records)
    cols = IDENTITY_COLUMNS.copy()
    for block in period_blocks:
        for day_date in block['dates']:
            cols.append(format_day_header_fr(day_date))
        cols.append(block['total_label'])
        if block.get('is_sunday'):
            cols.extend([block['amount_label_t1'], block['amount_label_t2']])
        cols.append(block['amount_label'])
    cols.extend([f'TOTAL {qty_label}', 'TOTAL A PAYER'])

    for col in cols:
        if col not in res_df.columns:
            res_df[col] = 0
            
    return res_df[cols].sort_values(by='TOTAL A PAYER', ascending=False)

def build_tariff_grid(engine, group_channels, data_type, date_ref, periods):
    table_col = "taux_gadd" if "Add" in data_type else "taux_ads"
    ch = ", ".join([f"'{c}'" for c in group_channels])
    
    sql = text(f"""
        SELECT periode_nom, MAX({table_col}) as val
        FROM commission_tarifs
        WHERE type_agent IN ({ch})
          AND :d BETWEEN date_debut AND date_fin
        GROUP BY periode_nom
    """)
    row = {period: 0 for period in periods}
    with engine.connect() as conn:
        res = conn.execute(sql, {"d": date_ref}).fetchall()
        for r in res:
            p = r[0]
            val = int(r[1])
            if p in row:
                row[p] = val
    return [row]

def write_sheet(wb, sheet_name, df_pivot, date_start, date_end, periods, data_type="GADD", engine=None, group_channels=None):
    ws = wb.create_sheet(sheet_name)
    if df_pivot.empty:
        ws.cell(1, 1, f"Pas de données pour {data_type}")
        return

    # 1. EN-TÊTE : Période en colonne 1
    ws.cell(1, 1, f"Du {format_date_fr(date_start)}").font = Font(bold=True, italic=True)
    ws.cell(2, 1, f"Au {format_date_fr(date_end)}").font = Font(bold=True, italic=True)

    # 2. MINI-TABLEAU TARIFAIRE (Commence à la colonne 2)
    t_headers = [f"PU {period}" for period in periods]
    
    # Récupération des données tarifs (uniquement si un groupe de canaux est fourni)
    if group_channels:
        tariffs = build_tariff_grid(engine, group_channels, data_type, date_end, periods)
        item = tariffs[0] if tariffs else {}
    else:
        item = {}  # Pas de grille tarifaire pour les feuilles combinées

    # Filtrage : On ne garde que les tarifs > 0
    active_pairs = [(header, period) for header, period in zip(t_headers, periods) if item.get(period, 0) > 0]
    
    for idx, (h, k) in enumerate(active_pairs, 2): # Colonne B (2) et suivantes
        # En-tête
        cell_h = ws.cell(1, idx, h)
        cell_h.font = SUB_HEADER_FONT
        cell_h.fill = SUB_HEADER_FILL
        cell_h.border = THIN_BORDER
        cell_h.alignment = Alignment(horizontal="center")
        
        # Valeur
        cell_v = ws.cell(2, idx, item.get(k, 0))
        cell_v.border = THIN_BORDER
        cell_v.number_format = CURRENCY_FMT

    # 3. EN-TÊTES DU TABLEAU DE DONNÉES (Ligne 3)
    start_data_row = 3

    # 3. EN-TÊTES DU TABLEAU DE DONNÉES
    headers = list(df_pivot.columns)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_data_row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Identification des types de colonnes
    currency_cols = [h for h in headers if h.startswith("Mnt") or h == "TOTAL A PAYER"]
    cur_indices = [headers.index(x)+1 for x in currency_cols]
    
    qty_cols = [h for h in headers if h not in IDENTITY_COLUMNS and h not in currency_cols]
    qty_indices = [headers.index(x)+1 for x in qty_cols]

    # 4. ÉCRITURE DES LIGNES
    current_row = start_data_row + 1
    for r_idx, row in enumerate(df_pivot.itertuples(index=False), current_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, val)
            
            if headers[c_idx-1] == 'MSISDN' and pd.notna(val) and str(val).strip() != "":
                try: cell.value = int(val)
                except: cell.value = val

            if headers[c_idx-1] == 'USER NAME':
                cell.value = str(val) if val else ""

            if val is None or pd.isna(val) or val == 0:
                if c_idx in cur_indices or c_idx in qty_indices:
                    cell.value = 0
                else:
                    cell.value = ""
                
            cell.border = THIN_BORDER
            
            if c_idx in cur_indices: cell.number_format = CURRENCY_FMT
            if c_idx in qty_indices: cell.number_format = '0'

    # 5. LIGNE DES TOTAUX
    last_row = len(df_pivot) + start_data_row
    t_row = last_row + 1
    ws.cell(t_row, 5, "TOTAL GLOBAL").font = Font(bold=True)
    ws.cell(t_row, 5).border = THIN_BORDER
    
    sum_cols = qty_indices + cur_indices
    for c in sum_cols:
        col_ltr = ws.cell(2, c).column_letter
        ws.cell(t_row, c, f"=SUM({col_ltr}{start_data_row+1}:{col_ltr}{last_row})")
        ws.cell(t_row, c).font = Font(bold=True)
        ws.cell(t_row, c).border = THIN_BORDER
        if c in cur_indices: ws.cell(t_row, c).number_format = CURRENCY_FMT

    auto_width(ws)
    
    # 6. GEL DES VOLETS (FREEZE PANES)
    # Ligne 1-2 : Tarifs, Ligne 3 : En-têtes, Colonne A : User Name
    # Geler au niveau de B4 permet de garder l'en-tête (Ligne 3) statique avec les dates, et le nom figé.
    ws.freeze_panes = "B4"

def resolve_date_range(engine, args):
    """Determine la plage de dates selon : args CLI > config.py > extractions outputs."""

    # 1. Arguments CLI explicites (prioritaires)
    if args.week:
        return (datetime.strptime(args.week[0], "%Y-%m-%d").date(),
                datetime.strptime(args.week[1], "%Y-%m-%d").date())
    if args.date:
        d = datetime.strptime(args.date, "%Y-%m-%d").date()
        return (d, d)

    # 2. Config BI (AUTO_MODE ou MANUAL)
    if AUTO_MODE:
        processing_day = date.today()
        end = get_latest_source_date(engine)
        if not end:
            print("Aucune donnee dans les tables journalieres.")
            sys.exit(1)

        current_source_range = get_processed_source_range(processing_day)
        last_mail_context = load_last_mail_context()

        if (
            current_source_range
            and last_mail_context
            and current_source_range["source_start_date"] == last_mail_context["source_start_date"]
            and current_source_range["source_end_date"] == last_mail_context["source_end_date"]
        ):
            start = last_mail_context["commission_start_date"]
            print(
                f"♻ Range source identique au dernier mail "
                f"({format_date_fr(current_source_range['source_start_date'])} -> "
                f"{format_date_fr(current_source_range['source_end_date'])}) : "
                f"mise a jour detectee, recalcul depuis {format_date_fr(start)}."
            )
        else:
            last_comm_date = find_last_commission_date()
            if last_comm_date:
                dyn_start = last_comm_date + timedelta(days=1)
                start = end if dyn_start > end else dyn_start
            else:
                # Fallback global si aucun fichier de commission précédent n'existe
                start = end - timedelta(days=AUTO_DAYS_RANGE - 1)

        if current_source_range:
            retrospective_start = current_source_range["source_start_date"]
            retrospective_retries = detect_retry_dates_in_range(engine, retrospective_start, end)
            retrospective_dates = [
                datetime.strptime(item["date"], "%Y-%m-%d").date()
                for item in retrospective_retries
                if item.get("date")
            ]
            if retrospective_dates:
                retry_start = min(retrospective_dates)
                if retry_start < start:
                    print(
                        f"⚠ Donnees anormales detectees dans le perimetre du mail courant, "
                        f"reprise forcee depuis {format_date_fr(retry_start)}."
                    )
                    start = retry_start

        daily_anchor_start = load_daily_recompute_anchor(processing_day)
        if daily_anchor_start and daily_anchor_start < start:
            print(
                f"♻ Recalcul du jour conserve depuis {format_date_fr(daily_anchor_start)} "
                f"pour couvrir toute correction du meme jour."
            )
            start = daily_anchor_start

        pending_retry_dates = load_pending_retry_dates(engine, end)
        if pending_retry_dates:
            retry_start = min(pending_retry_dates)
            if retry_start < start:
                print(
                    f"⚠ Donnee manquante enfin arrivee : reprise depuis "
                    f"{format_date_fr(retry_start)} pour integrer ce jour."
                )
                start = retry_start

        start = update_daily_recompute_anchor(start, end, processing_day)

        return (start, end)
    else:
        end = datetime.strptime(MANUAL_END_DATE, "%Y-%m-%d").date()
        start = end - timedelta(days=MANUAL_RANGE_DAYS - 1)
        return (start, end)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=None)
    parser.add_argument("--week", nargs=2, default=None)
    args = parser.parse_args()

    engine = make_engine(MYSQL_DATABASE)

    date_start, date_end = resolve_date_range(engine, args)

    print(f"=== Génération des Commissions (UX Mobile-First) ===")
    print(f"    Période : {format_date_fr(date_start)} -> {format_date_fr(date_end)}")

    df_gadd_raw = fetch_data(engine, "vw_commission_gadd", "gadd", "commission_gadd", date_start, date_end)
    df_ads_raw  = fetch_data(engine, "vw_commission_ads", "ads", "commission_ads", date_start, date_end)

    # Canal normalise (canonique) : robuste aux variantes 'BA AGENCE'/'BA_AGENCE', 'BA CLASSIQUE'/'BA', etc.
    if not df_gadd_raw.empty:
        df_gadd_raw['canal_norm'] = df_gadd_raw['real_channel'].map(canonical_channel)
    if not df_ads_raw.empty:
        df_ads_raw['canal_norm'] = df_ads_raw['real_channel'].map(canonical_channel)

    # Groupes du fichier exprimes en canaux CANONIQUES (= type_agent des tarifs).
    FILE_GROUPS = {
        "BA Animation": ["Animation Pick-up", "Animation POS"],
        "BA Classiques & BA AGENCE": ["BA", "BA_AGENCE"],
        "MA Acquisition": ["MA"]
    }
    all_group_channels = sorted({channel for channels in FILE_GROUPS.values() for channel in channels})
    active_periods = get_active_periods(engine, date_start, date_end, all_group_channels)
    day_dates = list(iter_dates(date_start, date_end))
    date_period_map = get_date_period_map(
        engine,
        date_start,
        date_end,
        all_group_channels,
        frames=[df_gadd_raw, df_ads_raw],
    )
    period_blocks = build_period_blocks(day_dates, date_period_map)

    OUTPUTS.mkdir(exist_ok=True)

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    # Combiner tous les groupes en 2 feuilles : GADD (tous agents) + ADS/New Users (tous agents)
    gadd_frames = []
    ads_frames  = []

    for group_name, group_channels in FILE_GROUPS.items():
        df_g_grp = df_gadd_raw[df_gadd_raw['canal_norm'].isin(group_channels)].copy() if not df_gadd_raw.empty else pd.DataFrame()
        df_a_grp = df_ads_raw[df_ads_raw['canal_norm'].isin(group_channels)].copy()  if not df_ads_raw.empty  else pd.DataFrame()

        if not df_g_grp.empty:
            gadd_frames.append(pivot_data(df_g_grp, day_dates, period_blocks, qty_label="Add"))
        if not df_a_grp.empty:
            ads_frames.append(pivot_data(df_a_grp, day_dates, period_blocks, qty_label="ADS"))

    df_all_gadd = (
        pd.concat(gadd_frames, ignore_index=True)
          .sort_values('TOTAL A PAYER', ascending=False)
          .reset_index(drop=True)
        if gadd_frames else pd.DataFrame()
    )
    df_all_ads = (
        pd.concat(ads_frames, ignore_index=True)
          .sort_values('TOTAL A PAYER', ascending=False)
          .reset_index(drop=True)
        if ads_frames else pd.DataFrame()
    )

    write_sheet(wb, "GADD", df_all_gadd, date_start, date_end, active_periods, "Tous - GADD",
                engine=engine, group_channels=None)
    write_sheet(wb, "ADS (New Users)", df_all_ads, date_start, date_end, active_periods, "Tous - ADS",
                engine=engine, group_channels=None)

    pending_retries = refresh_commission_state(engine, date_start, date_end)
    if pending_retries:
        retry_text = ", ".join(f"{item['metric']} {item['date']}" for item in pending_retries)
        print(f"\n⚠ Reprise automatique programmee pour : {retry_text}")

    # ── GARDE-FOU : NE JAMAIS COMMISSIONNER UNE PERIODE ENTIEREMENT A ZERO ──
    #
    # Cette verification ne depend PAS des jours voisins, contrairement a
    # `detect_pending_retries` : une periode commissionnee ou personne n'a rien
    # fait n'est jamais legitime a payer, qu'elle soit encadree ou non. C'est le
    # cas du 2026-08-12, ou la periode etait le seul 11/08 -- donc sans jour
    # precedent DANS la plage, donc invisible pour la detection par voisinage.
    #
    # La cause est en amont et n'est pas reparable ici : la colonne du jour
    # n'etait pas remplie dans le classeur source (`GLOBALE PERFORMANCE
    # 20260810.xlsx`). On refuse de produire le classeur plutot que d'en envoyer
    # un a zero. L'etat de reprise, lui, est DEJA persiste ci-dessus : la date
    # sera reprise d'elle-meme quand la donnee arrivera.
    totaux_periode = get_daily_metric_totals(engine, date_start, date_end)
    if totaux_periode and not any(has_any_activity(t) for t in totaux_periode.values()):
        jours = ", ".join(d.isoformat() for d in sorted(totaux_periode))

        # 🔴 INSCRIRE LA REPRISE ICI, ET NE PAS COMPTER SUR `refresh_commission_state`.
        #
        # J'ai d'abord ecrit que « l'etat de reprise est deja persiste ci-dessus,
        # la date se rattrapera d'elle-meme ». LE REJEU DU 2026-08-13 A PROUVE QUE
        # C'ETAIT FAUX : apres blocage sur le 11/08, `pending_retries` ne portait
        # que les deux vieilles dates de juin et juillet.
        #
        # La raison est structurelle. `detect_pending_retries` raisonne par
        # VOISINAGE, et la periode commissionnee etait le seul 11/08 : pas de
        # jour precedent DANS la plage, donc aucun voisin actif, donc rien a
        # detecter. C'est exactement la limite que le garde-fou ci-dessous existe
        # pour couvrir — mais couvrir l'ENVOI ne suffit pas : sans cette
        # inscription, la journee serait bloquee puis OUBLIEE, et il faudrait un
        # geste humain pour la rattraper le jour ou la donnee arrive.
        etat = load_commission_state()
        en_attente = list(etat.get("pending_retries", []))
        en_attente.extend({
            "date": d.isoformat(),
            "metric": "GADD+ADS",
            "reason": "periode entierement a 0 — commission bloquee, en attente de la donnee source",
        } for d in sorted(totaux_periode))
        # `save_commission_state` deduplique par (date, metric) : un blocage
        # repete n'empile pas de doublons.
        save_commission_state({"pending_retries": en_attente})

        print(
            "\n" + "=" * 60
            + "\n  BLOQUE : periode entierement a 0 (GADD et ADS)."
            + f"\n  Jour(s) concerne(s) : {jours}"
            + "\n  Aucun classeur n'est genere, aucun rapport n'est envoye."
            + "\n  Cause probable : colonne du jour non remplie dans le classeur"
            + "\n  source. Demander un renvoi ; la reprise est deja programmee."
            + "\n" + "=" * 60,
            file=sys.stderr,
        )
        sys.exit(5)

    print("\n✅ Feuilles ajoutées : GADD, ADS (New Users)")

    out_name = f"Commission LKA {date_start.strftime('%d-%m-%Y')} - {date_end.strftime('%d-%m-%Y')}.xlsx"
    out_path = OUTPUTS / out_name
    wb.save(out_path)

    current_source_range = save_last_mail_context(date_start, date_end)
    if current_source_range:
        print(
            "  [Contexte mail] Range source memorise : "
            f"{format_date_fr(current_source_range['source_start_date'])} -> "
            f"{format_date_fr(current_source_range['source_end_date'])}"
        )

    # ── Vérification des totaux ──
    verify_totals(
        engine, df_gadd_raw, df_ads_raw, df_all_gadd, df_all_ads,
        day_dates, date_start, date_end, all_group_channels
    )

    print(f"\n✅ Fichier généré : {out_path.name}")

if __name__ == "__main__":
    main()