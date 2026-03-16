import openpyxl
from openpyxl.utils import get_column_letter

file_path = r"d:\LKA\Perf_commissions\Point Financier Superviseur BA Dec 2025.xlsx"
sheets_to_extract = ['Fixe', 'Variable (Performance)', 'Variable 15 BA par Jour', 'Superviseur Mercenaire prim NA']

print("=" * 80)
wb = openpyxl.load_workbook(file_path, data_only=False)

for sheet_name in sheets_to_extract:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    print(f"\n[{sheet_name}]")
    
    # Headers (Row 1)
    headers = []
    for col in range(1, min(15, ws.max_column + 1)):
        val = str(ws.cell(1, col).value) if ws.cell(1, col).value else ""
        headers.append(val)
    print("HEADERS: " + " | ".join(headers))
    
    # First 4 Data Rows (Row 2-5)
    for row in range(2, 6):
        if row > ws.max_row: break
        row_data = []
        for col in range(1, min(15, ws.max_column + 1)):
            cell = ws.cell(row, col)
            val = f"={cell.value}" if cell.data_type == 'f' else str(cell.value)
            row_data.append(val[:35] if val else "")
        print(f"R{row} : " + " | ".join(row_data))
