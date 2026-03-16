"""
Script de vérification:
Compare les totaux exacts présents dans le vieux fichier Excel de Décembre 2025
avec la logique SQL qu'on vient d'écrire en mémoire.
"""
import openpyxl
import pandas as pd

# Données attendues tirées DIRECTEMENT du fichier 'Point Financier Superviseur BA Dec 2025.xlsx'
# Nous avions sorti : 
# Recap:
# D4=Fixe!G1, D5='Variable (Performance)'!J1, D6='Variable 15 BA par Jour'!AK2, D7='Superviseur Mercenaire prim NA'!G2
# Nous avions vu qu'Alberic etait R3, Amos R4. 

file_path = r"d:\LKA\Perf_commissions\Point Financier Superviseur BA Dec 2025.xlsx"
wb_v = openpyxl.load_workbook(file_path, data_only=True)

# 1. Extraction Totaux de l'Ancien Fichier
old_data = {}
ws_fixe = wb_v['Fixe']
for r in range(3, 8):
    sup = ws_fixe.cell(r, 2).value
    if sup: old_data[sup] = {'Fixe': ws_fixe.cell(r, 7).value or 0}

ws_var = wb_v['Variable (Performance)']
for r in range(3, 8):
    sup = ws_var.cell(r, 2).value
    if sup and sup in old_data:
        old_data[sup]['Variable'] = ws_var.cell(r, 10).value or 0
        
ws_15 = wb_v['Variable 15 BA par Jour']
for r in range(4, 9):
    sup = ws_15.cell(r, 1).value
    if sup and sup in old_data:
        old_data[sup]['15BA_Jour'] = ws_15.cell(r, 37).value or 0

ws_merc = wb_v['Superviseur Mercenaire prim NA']
for r in range(4, 8):
    sup = ws_merc.cell(r, 3).value
    if sup:
        if sup not in old_data: old_data[sup] = {'Fixe':0, 'Variable':0, '15BA_Jour':0}
        old_data[sup]['Mercenaire'] = ws_merc.cell(r, 7).value or 0

print("="*60)
print("1) ANCIENNES VALEURS (Extraites de Excel 'Décembre 2025')")
print("="*60)
for sup, primes in old_data.items():
    total_old = primes.get('Fixe',0) + primes.get('Variable',0) + primes.get('15BA_Jour',0) + primes.get('Mercenaire',0)
    print(f"[{sup}] Total Ancien: {total_old:,.0f} FCFA")
    print(f"    Détail -> Fixe:{primes.get('Fixe',0):.0f}, Var:{primes.get('Variable',0):.0f}, 15BA:{primes.get('15BA_Jour',0):.0f}, Merc:{primes.get('Mercenaire',0):.0f}")

print("\n" + "="*60)
print("2) NOUVELLES VALEURS (SQL et pandas)")
print("=> A comparer en exécutant python scripts/07_point_financier_superviseur.py")
print("="*60)
