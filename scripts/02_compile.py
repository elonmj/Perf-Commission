"""
scripts/02_compile.py — Étape 2 : fusionner les nouvelles données dans les fichiers cumulatifs.

Lit les outputs de 01_process.py (format long) et les fusionne dans les fichiers
cumulatifs (format wide : Username × dates en colonnes) :
  - cumul/GADD_cumul.xlsx
  - cumul/ADS_cumul.xlsx

Migration initiale : si les fichiers cumulatifs n'existent pas, copie les
fichiers existants (validbyuser010525.xlsx → GADD_cumul, ADS.xlsx → ADS_cumul).

Usage :
  py -3 scripts/02_compile.py
  py -3 scripts/02_compile.py --migrate   (migration initiale depuis les anciens fichiers)
"""

import sys
import os
import shutil
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT    = Path(__file__).parent.parent
OUTPUTS = ROOT / "outputs"
CUMUL   = ROOT / "cumul"

sys.path.insert(0, str(ROOT))

GADD_CUMUL = CUMUL / "GADD_cumul.xlsx"
ADS_CUMUL  = CUMUL / "ADS_cumul.xlsx"

# Anciens fichiers (migration initiale)
OLD_GADD = ROOT / "data" / "validbyuser010525.xlsx"
OLD_ADS  = ROOT / "data" / "ADS.xlsx"


def migrate_initial():
    """Copie les fichiers cumulatifs originaux dans cumul/ si pas encore fait."""
    CUMUL.mkdir(exist_ok=True)

    migrated = False
    if OLD_GADD.exists() and not GADD_CUMUL.exists():
        shutil.copy2(OLD_GADD, GADD_CUMUL)
        print(f"  Migration : {OLD_GADD.name} → {GADD_CUMUL.name}")
        migrated = True
    if OLD_ADS.exists() and not ADS_CUMUL.exists():
        shutil.copy2(OLD_ADS, ADS_CUMUL)
        print(f"  Migration : {OLD_ADS.name} → {ADS_CUMUL.name}")
        migrated = True

    if not migrated:
        print("  Aucune migration nécessaire (fichiers cumulatifs déjà présents).")
    return migrated


# ── Styles pour les fichiers cumulatifs ───────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def load_cumul_wide(path: Path) -> pd.DataFrame:
    """
    Charge un fichier cumulatif (format wide : Username, date1, date2, ...).
    Row 0 = header (Username, DD/MM/YYYY, DD/MM/YYYY, ...).
    """
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_excel(path, sheet_name=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def long_to_wide(df_long: pd.DataFrame, value_col: str) -> pd.DataFrame:
    """
    Convertit un DataFrame long (user_name, msisdn_momo, perf_date, value)
    en format wide (Username, DD/MM/YYYY, DD/MM/YYYY, ...).
    """
    if df_long.empty:
        return pd.DataFrame()

    df = df_long.copy()
    # Formater les dates en DD/MM/YYYY pour les colonnes
    df["date_str"] = pd.to_datetime(df["perf_date"]).dt.strftime("%d/%m/%Y")

    pivot = df.pivot_table(
        index="user_name",
        columns="date_str",
        values=value_col,
        aggfunc="first",
    ).reset_index()

    pivot.columns.name = None
    pivot = pivot.rename(columns={"user_name": "Username"})
    return pivot


def merge_wide(existing: pd.DataFrame, new_data: pd.DataFrame) -> pd.DataFrame:
    """
    Fusionne les nouvelles données (wide) dans le cumulatif existant (wide).
    Clé de merge : Username.
    Nouvelles dates ajoutées comme colonnes, valeurs mises à jour si déjà présentes.
    Nouveaux agents ajoutés en bas.
    """
    if existing.empty:
        return new_data
    if new_data.empty:
        return existing

    # Identifier les colonnes date du new_data (tout sauf Username)
    new_date_cols = [c for c in new_data.columns if c != "Username"]

    # Merge outer sur Username
    merged = existing.merge(new_data, on="Username", how="outer", suffixes=("", "_new"))

    # Pour chaque date du nouveau fichier, mettre à jour
    for col in new_date_cols:
        new_col = f"{col}_new"
        if new_col in merged.columns:
            # Écraser avec la nouvelle valeur si elle existe
            mask = merged[new_col].notna()
            if col in merged.columns:
                merged.loc[mask, col] = merged.loc[mask, new_col]
            else:
                merged[col] = merged[new_col]
            merged.drop(columns=[new_col], inplace=True)
        elif col not in merged.columns:
            merged[col] = new_data.set_index("Username")[col].reindex(merged["Username"]).values

    return merged


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Supprime les lignes avec Username invalide (vide, 0, numérique)."""
    if df.empty or "Username" not in df.columns:
        return df
    df = df[df["Username"].notna()].copy()
    df = df[df["Username"].astype(str).str.strip() != ""]
    df = df[~df["Username"].astype(str).str.strip().str.replace(".", "", regex=False)
            .str.replace("-", "", regex=False).str.replace("_", "", regex=False)
            .str.isdigit()]
    return df.reset_index(drop=True)


def validate_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Supprime les colonnes dont la date est dans le futur (ex: 2027)."""
    from datetime import date as dt_date
    today = dt_date.today()
    valid_cols = ["Username"]
    dropped = []
    for c in df.columns:
        if c == "Username":
            continue
        try:
            d = pd.to_datetime(c, format="%d/%m/%Y").date()
            if d <= today:
                valid_cols.append(c)
            else:
                dropped.append(c)
        except Exception:
            valid_cols.append(c)  # garder les colonnes non-date
    if dropped:
        print(f"  Colonnes supprimees (dates futures) : {dropped}")
    return df[[c for c in valid_cols if c in df.columns]]


def sort_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Trie les colonnes date (DD/MM/YYYY) chronologiquement."""
    non_date = ["Username"]
    date_cols = [c for c in df.columns if c not in non_date]

    # Parser et trier
    parsed = []
    for c in date_cols:
        try:
            d = pd.to_datetime(c, format="%d/%m/%Y")
            parsed.append((d, c))
        except Exception:
            parsed.append((pd.Timestamp.max, c))  # colonnes non-date à la fin

    parsed.sort(key=lambda x: x[0])
    sorted_cols = non_date + [c for _, c in parsed]

    return df[[c for c in sorted_cols if c in df.columns]]


def find_latest_output(prefix: str) -> Path | None:
    """Trouve le fichier le plus récent dans outputs/ avec le préfixe donné."""
    files = sorted(
        OUTPUTS.glob(f"{prefix}_*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    return files[0] if files else None


def styled_to_excel(df: pd.DataFrame, path: Path):
    """Sauvegarde un DataFrame en Excel avec entêtes colorées et colonnes auto-width."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Écrire les en-têtes
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(1, c_idx, col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Écrire les données
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                cell.value = None
            elif isinstance(val, str) and val.replace(".", "").replace("-", "").isdigit():
                try:
                    cell.value = int(float(val))
                except (ValueError, TypeError):
                    cell.value = val
            else:
                cell.value = val

    # Auto-width des colonnes
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            val_str = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 18)

    # Figer la première ligne et la première colonne
    ws.freeze_panes = "B2"

    wb.save(path)
    wb.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--migrate", action="store_true", help="Migration initiale")
    args = parser.parse_args()

    CUMUL.mkdir(exist_ok=True)

    print("=== 02_compile.py — Compilation cumulatifs ===\n")

    # ── Migration initiale si nécessaire ──────────────────────────────────────
    if args.migrate or not GADD_CUMUL.exists() or not ADS_CUMUL.exists():
        migrate_initial()
        print()

    # ── Charger outputs de 01_process.py ──────────────────────────────────────
    gadd_long_path = find_latest_output("gadd_long")
    ads_long_path  = find_latest_output("ads_long")

    if not gadd_long_path and not ads_long_path:
        print("  Aucun fichier à compiler dans outputs/.")
        print("  Lancez d'abord : py -3 scripts/01_process.py")
        sys.exit(1)

    # ── GADD ──────────────────────────────────────────────────────────────────
    if gadd_long_path:
        print(f"GADD source : {gadd_long_path.name}")
        df_gadd_long = pd.read_excel(gadd_long_path)
        new_gadd_wide = long_to_wide(df_gadd_long, "gadd")
        print(f"  Nouvelles dates GADD : {[c for c in new_gadd_wide.columns if c != 'Username']}")

        existing_gadd = load_cumul_wide(GADD_CUMUL)
        if not existing_gadd.empty:
            print(f"  Cumulatif existant : {len(existing_gadd)} agents, "
                  f"{len(existing_gadd.columns) - 1} dates")

        merged_gadd = merge_wide(existing_gadd, new_gadd_wide)
        merged_gadd = clean_dataframe(merged_gadd)
        merged_gadd = validate_date_columns(merged_gadd)
        merged_gadd = sort_date_columns(merged_gadd)

        styled_to_excel(merged_gadd, GADD_CUMUL)
        date_cols_g = [c for c in merged_gadd.columns if c != "Username"]
        print(f"  GADD cumulatif mis à jour : {len(merged_gadd)} agents, {len(date_cols_g)} dates")
        print(f"  → {GADD_CUMUL.name}")
    else:
        print("  Pas de nouvelles données GADD.")

    print()

    # ── ADS ───────────────────────────────────────────────────────────────────
    if ads_long_path:
        print(f"ADS source : {ads_long_path.name}")
        df_ads_long = pd.read_excel(ads_long_path)
        new_ads_wide = long_to_wide(df_ads_long, "ads")
        print(f"  Nouvelles dates ADS : {[c for c in new_ads_wide.columns if c != 'Username']}")

        existing_ads = load_cumul_wide(ADS_CUMUL)
        if not existing_ads.empty:
            print(f"  Cumulatif existant : {len(existing_ads)} agents, "
                  f"{len(existing_ads.columns) - 1} dates")

        merged_ads = merge_wide(existing_ads, new_ads_wide)
        merged_ads = clean_dataframe(merged_ads)
        merged_ads = validate_date_columns(merged_ads)
        merged_ads = sort_date_columns(merged_ads)

        styled_to_excel(merged_ads, ADS_CUMUL)
        date_cols_a = [c for c in merged_ads.columns if c != "Username"]
        print(f"  ADS cumulatif mis à jour : {len(merged_ads)} agents, {len(date_cols_a)} dates")
        print(f"  → {ADS_CUMUL.name}")
    else:
        print("  Pas de nouvelles données ADS.")

    print()
    print("Étape 2 terminée. Prochaine étape :")
    print("  - Premier chargement : py -3 scripts/03_upload.py")
    print("  - Re-run (sync)      : py -3 scripts/04_sync.py")


if __name__ == "__main__":
    main()
