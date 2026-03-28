import sys
import os
import json
import base64
import mimetypes
from pathlib import Path
from datetime import datetime, timedelta, date as date_cls

import pandas as pd
import numpy as np
from sqlalchemy import text
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from connections.config import MYSQL_DATABASE, TABLE_DAILY_GADD, RECIPIENTS_WEEKLY_BA_ACTIVITY
from connections.connect import make_engine
import importlib

fetch_mail = importlib.import_module("scripts.00_fetch_mail")
get_gmail_service = fetch_mail.get_gmail_service

TRACKER_FILE = ROOT / "data" / "processed_ba_activity_weeks.json"

def get_target_week(max_date, force_current=False):
    if force_current:
        days_since_sunday = (max_date.weekday() + 1) % 7
        start_date = max_date - timedelta(days=days_since_sunday)
        end_date = max_date
    else:
        days_since_saturday = (max_date.weekday() + 2) % 7
        last_saturday = max_date - timedelta(days=days_since_saturday)
        last_sunday = last_saturday - timedelta(days=6)
        start_date = last_sunday
        end_date = last_saturday
    return start_date, end_date

def load_processed_weeks():
    if not TRACKER_FILE.exists(): return []
    try: return json.loads(TRACKER_FILE.read_text(encoding="utf-8"))
    except: return []

def save_processed_week(week_id):
    TRACKER_FILE.parent.mkdir(exist_ok=True)
    weeks = load_processed_weeks()
    if week_id not in weeks:
        weeks.append(week_id)
        TRACKER_FILE.write_text(json.dumps(weeks), encoding="utf-8")

def send_email_with_attachment(service, subject, html_content, attachment_path, to_emails):
    if not to_emails:
        print("  [DEBUG] Aucun destinataire spécifié. Envoi d'email ignoré.")
        return

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = "me"
    msg["To"] = to_emails

    msg.add_alternative(html_content, subtype='html')

    path = Path(attachment_path)
    if path.exists():
        ctype, encoding = mimetypes.guess_type(str(path))
        if ctype is None or encoding is not None:
            ctype = 'application/octet-stream'
        maintype, subtype = ctype.split('/', 1)
        with open(path, 'rb') as f:
            msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=path.name)
            
    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    service.users().messages().send(userId="me", body={"raw": raw_message}).execute()

def format_excel_headers(ws, fill_color="E26B0A", font_color="FFFFFF"):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color=font_color)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    ws.freeze_panes = "A2"

def generate_ba_actif_weekly(engine, start_week, end_week):
    month_str = end_week.strftime('%Y-%m')
    start_of_month = pd.to_datetime(f"{month_str}-01").date()

    # Query with Region included and no generic "Username oui"
    query = f"""
    SELECT 
        a.user_name as `USER NAME`,
        a.dealer as `DEALER`,
        a.region as `REGION`,
        a.departement as `DEPARTEMENT`,
        a.commune as `COMMUNE`,
        a.enterprise_name as `ENTERPRISE_NAME`,
        a.agent_name as `AGENT_NAME`,
        a.momo_msisdn as `MOMO_MSISDN`,
        a.p2p_msisdn as `P2P MSISDN`,
        a.real_channel as `REAL_CHANNEL`,
        a.supervisor_full_name as `Superviseur`,
        a.sup_momo_msisdn as `SUP_MOMO_MSISDN`,
        a.tss_name as `TSS NAME`,
        a.other_dealers as `OTHER DEALERS`,
        a.numero_pulse as `Numero Pulse`,
        a.id_pulse as `Id pulse`,
        g.perf_date as `perf_date`,
        COALESCE(g.gadd, 0) as `gadd`
    FROM lka_client_mtn.lka_usernames a
    LEFT JOIN daily_gadd g ON a.user_name = g.user_name 
                          AND g.perf_date >= '{start_of_month}' 
                          AND g.perf_date <= '{end_week}'
    WHERE a.supervisor_full_name IS NOT NULL AND a.supervisor_full_name != '' AND a.supervisor_full_name != 'None'
    """
    df = pd.read_sql(query, engine)
    
    idx_cols = [
        'REGION', 'Superviseur', 'USER NAME', 'DEALER', 'DEPARTEMENT', 'COMMUNE', 
        'ENTERPRISE_NAME', 'AGENT_NAME', 'MOMO_MSISDN', 'P2P MSISDN', 
        'REAL_CHANNEL', 'SUP_MOMO_MSISDN', 'TSS NAME', 
        'OTHER DEALERS', 'Numero Pulse', 'Id pulse'
    ]
    
    pivot_df = df.pivot_table(index=idx_cols, columns='perf_date', values='gadd', aggfunc='sum')
    pivot_df = pivot_df.reset_index()
    pivot_df.columns = [c.strftime('%Y-%m-%d') if isinstance(c, (pd.Timestamp, date_cls)) else c for c in pivot_df.columns]
        
    date_range = pd.date_range(start=start_of_month, end=end_week)
    for dt in date_range:
        date_col = dt.date().strftime('%Y-%m-%d')
        if date_col not in pivot_df.columns:
            pivot_df[date_col] = 0
            
    date_cols = [c for c in list(pivot_df.columns) if isinstance(c, str) and len(c) == 10 and c[4] == '-' and c[7] == '-']
    date_cols_sorted = sorted(date_cols)

    pivot_df = pivot_df[idx_cols + date_cols_sorted].fillna(0).copy()
    for col in date_cols_sorted:
        pivot_df[col] = pd.to_numeric(pivot_df[col], errors='coerce').fillna(0)
    
    month_dates = [c for c in date_cols_sorted]
    last_week_dates = [c for c in date_cols_sorted if pd.to_datetime(c).date() >= start_week and pd.to_datetime(c).date() <= end_week]
    
    pivot_df['Actif pour le mois'] = (pivot_df[month_dates].sum(axis=1) > 0).astype(int)
    pivot_df['Actif La semaine denier'] = (pivot_df[last_week_dates].sum(axis=1) > 0).astype(int)
    
    recap = pivot_df.groupby(['REGION', 'Superviseur']).agg(
        Nb_des_BA=('USER NAME', 'count'),
        Nb_Actif_pour_le_mois=('Actif pour le mois', 'sum'),
        Nb_Actif_La_semaine_denier=('Actif La semaine denier', 'sum')
    ).reset_index()
    
    recap.rename(columns={
        'Superviseur': 'Description', 
        'Nb_des_BA': 'Nb des BA', 
        'Nb_Actif_pour_le_mois': 'Nb Actif pour le mois', 
        'Nb_Actif_La_semaine_denier': 'Nb Actif La semaine denier'
    }, inplace=True)
                           
    # PCTs
    recap['PCT Actif le mois'] = recap['Nb Actif pour le mois'] / recap.replace(0, pd.NA)['Nb des BA']
    recap['PCT Actif la semaine dernier'] = recap['Nb Actif La semaine denier'] / recap.replace(0, pd.NA)['Nb des BA']
    recap['PCT Actif le mois'] = recap['PCT Actif le mois'].fillna(0)
    recap['PCT Actif la semaine dernier'] = recap['PCT Actif la semaine dernier'].fillna(0)

    # Calculer le score de performance par région (moyenne du PCT Actif la semaine dernière)
    region_scores = recap.groupby('REGION')['PCT Actif la semaine dernier'].mean().reset_index()
    region_scores.rename(columns={'PCT Actif la semaine dernier': 'REGION_SCORE'}, inplace=True)
    
    # Fusionner pour trier : Régions les plus critiques (SCORE faible) d'abord, puis superviseurs critiques
    recap = recap.merge(region_scores, on='REGION')
    recap = recap.sort_values(by=['REGION_SCORE', 'REGION', 'PCT Actif la semaine dernier'], ascending=[True, True, True])
    
    # Garder le détail pivoté cohérent
    pivot_df = pivot_df.sort_values(by=['REGION', 'Superviseur'])

    output_dir = os.path.join(ROOT, "outputs")
    os.makedirs(output_dir, exist_ok=True)
    out_file = os.path.join(output_dir, f"Weekly_BA_Actif_par_Superviseur_{start_week}_to_{end_week}.xlsx")
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        # On n'utilise pas recap.to_excel directement pour pouvoir insérer des lignes vides
        workbook = writer.book
        ws_recap = workbook.create_sheet('Recap')
        
        # En-têtes
        cols = list(recap.columns)
        # Supprimer le score technique avant l'export
        if 'REGION_SCORE' in cols: cols.remove('REGION_SCORE')
        
        for c_idx, col_name in enumerate(cols, 1):
            cell = ws_recap.cell(row=1, column=c_idx, value=col_name)
        
        current_row = 2
        last_region = None
        
        for _, row_data in recap.iterrows():
            region = row_data['REGION']
            # Insérer une ligne vide au changement de région (sauf la première fois)
            if last_region is not None and region != last_region:
                current_row += 1
            
            for c_idx, col_name in enumerate(cols, 1):
                val = row_data[col_name]
                cell = ws_recap.cell(row=current_row, column=c_idx, value=val)
                # Format Percentage pour les colonnes PCT
                if 'PCT' in col_name:
                    cell.number_format = '0.00%'
            
            last_region = region
            current_row += 1

        # Export de l'autre feuille normalement
        pivot_df.to_excel(writer, sheet_name='New Add', index=False)
        
        # Formatage final
        format_excel_headers(ws_recap, fill_color="4F81BD")
        ws_new_add = writer.sheets['New Add']
        format_excel_headers(ws_new_add, fill_color="4F81BD")

    return out_file

def main():
    print("=== 07_weekly_ba_activity.py ===")
    engine = make_engine(MYSQL_DATABASE)

    with engine.connect() as conn:
        row = conn.execute(text(f"SELECT MAX(perf_date) FROM {TABLE_DAILY_GADD}")).fetchone()
        
    if not row or not row[0]: return

    max_date = row[0]
    is_force = os.environ.get("FORCE_SEND_ALL") == "1"
    start_date, end_date = get_target_week(max_date, is_force)
    week_id = f"{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}"
    
    processed_weeks = load_processed_weeks()
    if week_id in processed_weeks and not is_force:
        print(f"  Semaine {week_id} déjà traitée (BA Actif).")
        sys.exit(0)

    ba_file = generate_ba_actif_weekly(engine, start_date, end_date)
    
    service = get_gmail_service()
    html_ba = f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #333;">
        <h2 style="color: #4F81BD;">Activité Hebdomadaire des BA par Superviseur</h2>
        <p><strong>Période :</strong> Semaine du {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}</p>
        <p>Bonjour,</p>
        <p>Veuillez trouver ci-joint le bilan d'activité (BA Actifs) regroupé par Région et Superviseur, classé par les meilleures performances.</p>
        <p>L'onglet <strong>Recap</strong> donne vos KPIs globaux, et l'onglet <strong>New Add</strong> contient le détail journalier.</p>
        <br>
        <p style="font-size: 12px; color: #666;"><em>Service Reporting Automatisé</em></p>
      </body>
    </html>
    """
    subject_ba = f"📈 BA Actifs par Superviseur/Région ({start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m')})"
    send_email_with_attachment(service, subject_ba, html_ba, ba_file, RECIPIENTS_WEEKLY_BA_ACTIVITY)

    save_processed_week(week_id)
    print("  Email Weekly BA Actif envoyé !")

if __name__ == "__main__":
    main()