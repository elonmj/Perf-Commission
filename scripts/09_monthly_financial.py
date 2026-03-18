import sys
import os
import json
import base64
import mimetypes
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
from sqlalchemy import text
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from connections.config import MYSQL_DATABASE, RECIPIENTS_MONTHLY_FINANCIAL, TABLE_DAILY_GADD
from connections.connect import make_engine
import importlib

fetch_mail = importlib.import_module("scripts.00_fetch_mail")
get_gmail_service = fetch_mail.get_gmail_service

TRACKER_FILE = ROOT / "data" / "processed_monthly_financial.json"

def load_processed_months():
    if not TRACKER_FILE.exists(): return []
    try: return json.loads(TRACKER_FILE.read_text(encoding="utf-8"))
    except: return []

def save_processed_month(month_id):
    TRACKER_FILE.parent.mkdir(exist_ok=True)
    months = load_processed_months()
    if month_id not in months:
        months.append(month_id)
        TRACKER_FILE.write_text(json.dumps(months), encoding="utf-8")

def send_email_with_attachment(service, subject, html_content, attachment_path, to_emails):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = "me"
    msg["To"] = to_emails

    msg.add_alternative(html_content, subtype='html')

    path = Path(attachment_path)
    if path.exists():
        ctype, encoding = mimetypes.guess_type(str(path))
        if ctype is None or encoding is not None:
            ctype = 'application/octet-stream'
        maintype, subtype = ctype.split('/', 1)
        with open(path, 'rb') as f:
            msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=path.name)
            
    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    service.users().messages().send(userId="me", body={"raw": raw_message}).execute()

def list_available_months(engine):
    query = "SELECT DISTINCT perf_month FROM vw_commission_superviseur ORDER BY perf_month DESC"
    res = pd.read_sql(query, engine)
    return res['perf_month'].tolist()

def get_last_complete_month(engine):
    # A complete month is usually the prior month relative to max date
    with engine.connect() as conn:
        row = conn.execute(text(f"SELECT MAX(perf_date) FROM {TABLE_DAILY_GADD}")).fetchone()
    if not row or not row[0]: return None
    
    max_date = row[0]
    # We define a "complete" month as a month that is before the month of max_date
    # Or, if today is late enough in the month, we can report on the immediate past month
    # But usually, it's just `max_date.replace(day=1) - timedelta(days=1)`.
    prior_month_date = max_date.replace(day=1) - timedelta(days=1)
    return prior_month_date.strftime('%Y-%m')

def generate_monthly_point_financier(engine, target_month):
    query_synth = f"SELECT * FROM vw_commission_superviseur WHERE perf_month = '{target_month}'"
    df_synth = pd.read_sql(query_synth, engine)
    if df_synth.empty: return None
        
    df_synth['TOTAL COMMISSION'] = (
        df_synth['prime_fixe'] + df_synth['prime_variable'] + 
        df_synth['prime_15ba_jour'] + df_synth['prime_mercenaire']
    )

    # Sort descending by GADD or Commission
    df_synth = df_synth.sort_values(by="TOTAL COMMISSION", ascending=False)

    query_detail = f"""
    SELECT 
        a.superviseur, a.user_name as agent, a.real_channel as type_agent,
        SUM(g.gadd) as total_gadd,
        COUNT(DISTINCT CASE WHEN g.gadd > 0 THEN g.perf_date END) as jours_actifs
    FROM agent_perf_info a
    JOIN daily_gadd g ON a.user_name = g.user_name
    WHERE DATE_FORMAT(g.perf_date, '%%Y-%%m') = '{target_month}'
    AND a.superviseur IS NOT NULL AND a.superviseur != ''
    GROUP BY a.superviseur, a.user_name, a.real_channel
    ORDER BY a.superviseur, total_gadd DESC
    """
    df_detail = pd.read_sql(query_detail, engine)
    
    output_dir = os.path.join(ROOT, "outputs")
    os.makedirs(output_dir, exist_ok=True)
    out_file = os.path.join(output_dir, f"Monthly_Point_Financier_{target_month}.xlsx")
    
    wb = Workbook()
    ws_synth = wb.active
    ws_synth.title = "Synthèse Financière"
    
    rename_cols = {
        'superviseur_name': 'Superviseur', 'type_superviseur': 'Catégorie', 'perf_month': 'Mois',
        'nb_ba_total': 'Nb BA Assignés', 'target_mensuel': 'Objectif Mensuel', 'total_new_add': 'GADD Total',
        'jours_actifs_15': 'Jours (>=15 BA Actifs)', 'prime_fixe': 'Prime Fixe', 'prime_variable': 'Prime Variable',
        'prime_15ba_jour': 'Prime 15 BA/Jour', 'prime_mercenaire': 'Prime Mercenaire', 'TOTAL COMMISSION': 'TOTAL À PAYER'
    }
    cols_to_keep = [c for c in rename_cols.keys() if c in df_synth.columns]
    df_synth_display = df_synth[cols_to_keep].rename(columns=rename_cols).copy()
    
    for r in dataframe_to_rows(df_synth_display, index=False, header=True): ws_synth.append(r)
        
    header_fill_1 = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    for cell in ws_synth[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill_1
        cell.alignment = Alignment(horizontal="center")
    for col in ws_synth.columns:
        ws_synth.column_dimensions[col[0].column_letter].width = 18
        
    ws_detail = wb.create_sheet(title="Détails Agents")
    df_detail_display = df_detail.rename(columns={
        'superviseur': 'Superviseur', 'agent': 'Nom Agent (BA)', 'type_agent': 'Type Agent', 
        'total_gadd': 'GADD Réalisé', 'jours_actifs': 'Jours Actifs (>0 GADD)'
    })
    for r in dataframe_to_rows(df_detail_display, index=False, header=True): ws_detail.append(r)
        
    header_fill_2 = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for cell in ws_detail[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill_2
    for col in ws_detail.columns:
        ws_detail.column_dimensions[col[0].column_letter].width = 22

    ws_synth.freeze_panes = "A2"
    ws_detail.freeze_panes = "A2"
    wb.save(out_file)
    return out_file


def main():
    print("=== 09_monthly_financial.py ===")
    engine = make_engine(MYSQL_DATABASE)

    target_month = get_last_complete_month(engine)
    if not target_month:
        print("  Aucun mois complet détecté.")
        sys.exit(0)

    processed_months = load_processed_months()
    if target_month in processed_months:
        print(f"  Mois {target_month} déjà traité (Financier).")
        sys.exit(0)

    pf_file = generate_monthly_point_financier(engine, target_month)
    if not pf_file:
        print(f"  Aucune donnée pour générer le Point Financier de {target_month}.")
        sys.exit(0)
    
    service = get_gmail_service()
    html_pf = f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #333;">
        <h2 style="color: #1F497D;">Point Financier Mensuel Superviseur - {target_month}</h2>
        <p>Le mois <strong>{target_month}</strong> est désormais clôturé.</p>
        <p>Bonjour,</p>
        <p>Veuillez trouver ci-joint le calcul définitif des primes financières (fixe, variable, etc.) pour chaque superviseur concernant le mois de {target_month}.</p>
        <p>Il regroupe le total financier (classé par performance décroissante) ainsi qu'un onglet contenant les détails par agent (BA).</p>
        <br>
        <p style="font-size: 12px; color: #666;"><em>Service Reporting Automatisé</em></p>
      </body>
    </html>
    """
    subject_pf = f"💰 Point Financier Définitif Superviseur ({target_month})"
    send_email_with_attachment(service, subject_pf, html_pf, pf_file, RECIPIENTS_MONTHLY_FINANCIAL)

    save_processed_month(target_month)
    print(f"  Email Monthly Financial ({target_month}) envoyé !")

if __name__ == "__main__":
    main()