"""
scripts/01_process.py — Étape 1 : traitement du fichier GLOBALE PERFORMANCE.

Lit le fichier Excel brut (multi-row headers), extrait les infos agent et les
métriques GADD/ADS par date, normalise, et produit :
  - outputs/gadd_long_YYYY-MM-DD.xlsx  (format long : msisdn_momo, date, gadd)
  - outputs/ads_long_YYYY-MM-DD.xlsx   (format long : msisdn_momo, date, ads)
  - outputs/agent_info_YYYY-MM-DD.xlsx (snapshot agent)
  - reports/quality_report_YYYY-MM-DD_HHMMSS.txt

Usage :
  py -3 scripts/01_process.py
  py -3 scripts/01_process.py --file "inputs/GLOBALE PERFORMANCE 20260310.xlsx"
"""

import sys
import os
import argparse
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

ROOT    = Path(__file__).parent.parent
INPUTS  = ROOT / "inputs"
OUTPUTS = ROOT / "outputs"
REPORTS = ROOT / "reports"

sys.path.insert(0, str(ROOT))
from connections.config import (
    PERF_HEADER_ROW, MSISDN_COLS, REGION_NORMALIZATION,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def find_latest_input(override: str = None) -> Path:
    if override:
        p = Path(override)
        if not p.exists():
            p = ROOT / override
        if not p.exists():
            raise FileNotFoundError(f"Fichier introuvable : {override}")
        return p
    files = sorted(
        [f for f in INPUTS.glob("*.xlsx") if not f.name.startswith("~$")],
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    if not files:
        raise FileNotFoundError(f"Aucun fichier .xlsx dans {INPUTS}")
    return files[0]


def parse_perf_file(path: Path):
    """
    Parse le fichier GLOBALE PERFORMANCE avec openpyxl (multi-row headers).

    Structure attendue (1-indexed) :
      Row 4  : dates (merged : col P-Q = date1, R-S = date2, ...)
      Row 5  : col E-O = noms agent cols, col P+ = GADD, ADS, GADD, ADS, ...
      Row 6+ : données

    Returns:
        df_agent : DataFrame agent info (user_name, region, ..., msisdn_momo)
        df_gadd  : DataFrame long (user_name, msisdn_momo, perf_date, gadd)
        df_ads   : DataFrame long (user_name, msisdn_momo, perf_date, ads)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    
    # Prise en charge dynamique de l'onglet : prendre le premier onglet, peu importe son nom
    if not wb.sheetnames:
        raise ValueError(f"Aucun onglet trouvé dans {path}")
    
    first_sheet_name = wb.sheetnames[0]
    ws = wb[first_sheet_name]
    print(f"  [Info] Lecture de l'onglet : '{first_sheet_name}'")

    max_row = ws.max_row
    max_col = ws.max_column

    # -- RECHERCHE DYNAMIQUE DES COLONNES --
    # Sonde ±2 lignes autour de PERF_HEADER_ROW — absorbe un décalage si des lignes
    # sont ajoutées en haut du fichier. Comparaison case-insensitive.
    real_header_row      = PERF_HEADER_ROW  # sera mis à jour si trouvé ailleurs
    real_agent_col_start = 0               # sentinel : 0 = non trouvé
    scan_start = max(1, PERF_HEADER_ROW - 2)
    scan_end   = PERF_HEADER_ROW + 2
    for probe_row in range(scan_start, scan_end + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(probe_row, c).value
            if v is not None and str(v).strip().lower() == "user_name":
                real_header_row      = probe_row
                real_agent_col_start = c
                break
        if real_agent_col_start != 0:
            break

    if real_agent_col_start == 0:
        raise ValueError(
            f"Colonne 'user_name' introuvable dans les lignes {scan_start}-{scan_end} "
            f"de '{path.name}'. Vérifiez la structure du fichier "
            f"(en-tête attendu autour de la ligne {PERF_HEADER_ROW})."
        )

    # Lignes date et données : immédiatement au-dessus / en-dessous du header détecté
    real_date_row       = real_header_row - 1  # ligne merged des dates
    real_data_start_row = real_header_row + 1  # première ligne de données

    # Recherche de la première colonne GADD ou ADS (case-insensitive)
    real_metric_col_start = 0  # sentinel : 0 = non trouvé
    for c in range(real_agent_col_start + 1, max_col + 1):
        val = ws.cell(real_header_row, c).value
        if val is not None and str(val).strip().upper() in ("GADD", "ADS"):
            real_metric_col_start = c
            break

    if real_metric_col_start == 0:
        raise ValueError(
            f"Colonne GADD/ADS introuvable après 'user_name' (col {real_agent_col_start}, "
            f"row {real_header_row}) dans '{path.name}'. Vérifiez la structure du fichier."
        )

    real_agent_col_end = real_metric_col_start - 1
    print(
        f"  [Info] Header row={real_header_row}, "
        f"agent cols={real_agent_col_start}-{real_agent_col_end}, "
        f"métriques col={real_metric_col_start}+, "
        f"date row={real_date_row}, données row={real_data_start_row}+"
    )

    DYNAMIC_AGENT_COLUMNS = []
    for c in range(real_agent_col_start, real_agent_col_end + 1):
        v = ws.cell(real_header_row, c).value
        if v is not None:
            v_str = str(v).strip().lower()
            if v_str == "tsa":
                v_str = "tss"
            DYNAMIC_AGENT_COLUMNS.append(v_str)
        else:
            DYNAMIC_AGENT_COLUMNS.append(f"col_{c}")

    # ── 1. Lire les dates (ligne merged au-dessus du header — seul le coin haut-gauche a la valeur)
    dates_by_col = {}
    for c in range(real_metric_col_start, max_col + 1):
        val = ws.cell(real_date_row, c).value
        if val is not None:
            if isinstance(val, datetime):
                dates_by_col[c] = val.date()
            else:
                try:
                    dates_by_col[c] = pd.to_datetime(str(val)).date()
                except Exception:
                    pass

    # Propager les dates aux colonnes ADS (merged → col suivante sans valeur)
    date_map = {}  # col_index → date
    last_date = None
    for c in range(real_metric_col_start, max_col + 1):
        if c in dates_by_col:
            last_date = dates_by_col[c]
        if last_date is not None:
            date_map[c] = last_date

    # ── 2. Lire les sub-headers (GADD/ADS) depuis le header row
    sub_headers = {}
    for c in range(real_metric_col_start, max_col + 1):
        val = ws.cell(real_header_row, c).value
        if val is not None:
            sub_headers[c] = str(val).strip().upper()

    # ── 3. Lire les données
    agent_rows = []
    gadd_records = []
    ads_records = []

    for r in range(real_data_start_row, max_row + 1):
        # Agent info (cols 5..15)
        agent_vals = []
        for c in range(real_agent_col_start, real_agent_col_end + 1):
            agent_vals.append(ws.cell(r, c).value)

        # Skip empty or invalid rows (None, empty, numeric-only like 0)
        user_name = agent_vals[0]
        if user_name is None or str(user_name).strip() == "":
            continue
        uname_str = str(user_name).strip()
        if uname_str.replace(".", "").replace("-", "").replace("_", "").isdigit():
            continue  # Skip numeric-only usernames (e.g. 0)

        agent_dict = dict(zip(DYNAMIC_AGENT_COLUMNS, agent_vals))
        agent_rows.append(agent_dict)

        msisdn = agent_dict.get("msisdn_momo")
        uname = uname_str

        # Metric cols (GADD, ADS alternating)
        for c in range(real_metric_col_start, max_col + 1):
            if c not in date_map or c not in sub_headers:
                continue
            val = ws.cell(r, c).value
            try:
                val = int(val) if val is not None else 0
            except (ValueError, TypeError):
                val = 0

            metric = sub_headers[c]
            d = date_map[c]

            if metric == "GADD":
                gadd_records.append({
                    "user_name": uname,
                    "msisdn_momo": msisdn,
                    "perf_date": d,
                    "gadd": val,
                })
            elif metric == "ADS":
                ads_records.append({
                    "user_name": uname,
                    "msisdn_momo": msisdn,
                    "perf_date": d,
                    "ads": val,
                })

    wb.close()

    df_agent = pd.DataFrame(agent_rows)
    df_gadd  = pd.DataFrame(gadd_records)
    df_ads   = pd.DataFrame(ads_records)

    return df_agent, df_gadd, df_ads


def normalize_regions(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    if "region" not in df.columns:
        return df, {}
    # Strip whitespace first
    df["region"] = df["region"].astype(str).str.strip()
    corrections = {}
    for raw, canonical in REGION_NORMALIZATION.items():
        mask = df["region"].astype(str).str.upper().str.strip() == raw.upper()
        count = mask.sum()
        if count > 0:
            df.loc[mask, "region"] = canonical
            corrections[raw] = {"canonical": canonical, "count": int(count)}
    return df, corrections


def convert_msisdn(df: pd.DataFrame) -> pd.DataFrame:
    for col in MSISDN_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
    return df


def build_quality_report(
    source: Path, n_agents: int, n_dates: int,
    gadd_rows: int, ads_rows: int,
    region_corrections: dict, dates_found: list,
) -> str:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "RAPPORT QUALITÉ — Perf_commissions",
        f"Généré le    : {ts}",
        f"Source       : {source.name}",
        "=" * 60,
        "",
        f"AGENTS           : {n_agents}",
        f"DATES TROUVÉES   : {n_dates}",
        f"  Dates : {', '.join(str(d) for d in sorted(dates_found))}",
        f"LIGNES GADD      : {gadd_rows}",
        f"LIGNES ADS       : {ads_rows}",
        "",
        "NORMALISATION DES RÉGIONS :",
    ]
    if region_corrections:
        for raw, info in region_corrections.items():
            lines.append(f"  '{raw}' → '{info['canonical']}' ({info['count']} corrigées)")
    else:
        lines.append("  Aucune correction nécessaire.")
    lines.append("")
    lines.append("=" * 60)
    lines.append("FIN DU RAPPORT")
    return "\n".join(lines)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default=None, help="Fichier GLOBALE PERFORMANCE (optionnel)")
    args = parser.parse_args()

    source = find_latest_input(args.file)
    print(f"Source : {source}")

    # ── 1. Parse ──────────────────────────────────────────────────────────────
    print("Lecture et extraction en cours...")
    df_agent, df_gadd, df_ads = parse_perf_file(source)
    print(f"  Agents : {len(df_agent)}")
    print(f"  GADD records : {len(df_gadd)}")
    print(f"  ADS records  : {len(df_ads)}")

    # ── 2. Normalisation agent ────────────────────────────────────────────────
    df_agent = convert_msisdn(df_agent)
    df_agent, region_corrections = normalize_regions(df_agent)

    if region_corrections:
        for raw, info in region_corrections.items():
            print(f"  Région corrigée : '{raw}' → '{info['canonical']}' ({info['count']})")

    # ── 3. Propager msisdn nettoyé dans gadd/ads ─────────────────────────────
    for col in MSISDN_COLS:
        if col in df_gadd.columns:
            df_gadd[col] = pd.to_numeric(df_gadd[col], errors="coerce").astype("Int64")
        if col in df_ads.columns:
            df_ads[col] = pd.to_numeric(df_ads[col], errors="coerce").astype("Int64")

    # ── 4. Dates extraites ────────────────────────────────────────────────────
    dates_gadd = sorted(df_gadd["perf_date"].unique()) if len(df_gadd) else []
    dates_ads  = sorted(df_ads["perf_date"].unique()) if len(df_ads) else []
    all_dates  = sorted(set(list(dates_gadd) + list(dates_ads)))
    print(f"  Dates : {all_dates}")

    # ── 5. Rapport qualité ────────────────────────────────────────────────────
    REPORTS.mkdir(exist_ok=True)
    report_ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    report_path = REPORTS / f"quality_report_{report_ts}.txt"
    report_text = build_quality_report(
        source, len(df_agent), len(all_dates),
        len(df_gadd), len(df_ads),
        region_corrections, all_dates,
    )
    report_path.write_text(report_text, encoding="utf-8")
    print(f"  Rapport → {report_path.name}")

    # ── 6. Export outputs ─────────────────────────────────────────────────────
    OUTPUTS.mkdir(exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")

    # Agent info
    agent_path = OUTPUTS / f"agent_info_{today}.xlsx"
    df_agent_export = df_agent.copy()
    for col in MSISDN_COLS:
        if col in df_agent_export.columns:
            df_agent_export[col] = df_agent_export[col].astype("float64")
    df_agent_export.to_excel(agent_path, index=False, sheet_name="Agents")
    print(f"  Agent info  → {agent_path.name}")

    # GADD long
    gadd_path = OUTPUTS / f"gadd_long_{today}.xlsx"
    df_gadd_export = df_gadd.copy()
    if "msisdn_momo" in df_gadd_export.columns:
        df_gadd_export["msisdn_momo"] = df_gadd_export["msisdn_momo"].astype("float64")
    df_gadd_export.to_excel(gadd_path, index=False, sheet_name="GADD")
    print(f"  GADD long   → {gadd_path.name}")

    # ADS long
    ads_path = OUTPUTS / f"ads_long_{today}.xlsx"
    df_ads_export = df_ads.copy()
    if "msisdn_momo" in df_ads_export.columns:
        df_ads_export["msisdn_momo"] = df_ads_export["msisdn_momo"].astype("float64")
    df_ads_export.to_excel(ads_path, index=False, sheet_name="ADS")
    print(f"  ADS long    → {ads_path.name}")

    # ── Résumé ────────────────────────────────────────────────────────────────
    print()
    if "region" in df_agent.columns:
        print("  ── Régions ──")
        for val, cnt in df_agent["region"].value_counts().items():
            print(f"     {val}: {cnt}")

    print()
    print("Étape 1 terminée. Prochaine étape :")
    print("  py -3 scripts/02_compile.py")


if __name__ == "__main__":
    main()
