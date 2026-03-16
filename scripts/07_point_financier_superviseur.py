import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
from connections.connect import make_engine

def list_available_months(engine):
    query = "SELECT DISTINCT perf_month FROM vw_commission_superviseur ORDER BY perf_month DESC"
    res = pd.read_sql(query, engine)
    return res['perf_month'].tolist()

def generate_report(target_month="2025-12"):
    engine = make_engine()
    
    # Vérification de l'existence des données
    months = list_available_months(engine)
    if target_month not in months:
        print(f"Attention: Le mois {target_month} n'est pas (encore) agrégé dans la base.")
        if months:
            print(f"Mois disponibles : {', '.join(months)}")
            target_month = months[0]
            print(f"Génération forcée sur le dernier mois disponible : {target_month}")
        else:
            return
            
    # 1. Extraction des données Synthèse
    query_synth = f"SELECT * FROM vw_commission_superviseur WHERE perf_month = '{target_month}'"
    df_synth = pd.read_sql(query_synth, engine)
    
    if df_synth.empty:
        print(f"Aucune erreur, mais 0 lignes pour {target_month}.")
        return
        
    # Calcul du Total dans Pandas (ajout de la colonne)
    df_synth['TOTAL COMMISSION'] = (
        df_synth['prime_fixe'] + 
        df_synth['prime_variable'] + 
        df_synth['prime_15ba_jour'] + 
        df_synth['prime_mercenaire']
    )
    
    # 2. Extraction des Détails par BA (la preuve)
    query_detail = f"""
    SELECT 
        a.superviseur, 
        a.user_name as agent, 
        a.real_channel as type_agent,
        SUM(g.gadd) as total_gadd,
        COUNT(DISTINCT CASE WHEN g.gadd > 0 THEN g.perf_date END) as jours_actifs
    FROM agent_perf_info a
    JOIN daily_gadd g ON a.user_name = g.user_name
    WHERE DATE_FORMAT(g.perf_date, '%%Y-%%m') = '{target_month}'
    AND a.superviseur IS NOT NULL AND a.superviseur != ''
    GROUP BY a.superviseur, a.user_name, a.real_channel
    ORDER BY a.superviseur, total_gadd DESC
    """
    df_detail = pd.read_sql(query_detail, engine)
    
    # 3. Création du Fichier Excel
    output_dir = os.path.join(ROOT, "outputs")
    os.makedirs(output_dir, exist_ok=True)
    out_file = os.path.join(output_dir, f"Point_Financier_Superviseur_{target_month}.xlsx")
    
    wb = Workbook()
    
    # ========= ONGLET 1 : SYNTHÈSE =========
    ws_synth = wb.active
    ws_synth.title = "Synthèse Financière"
    
    rename_cols = {
        'superviseur_name': 'Superviseur',
        'type_superviseur': 'Catégorie',
        'perf_month': 'Mois',
        'nb_ba_total': 'Nb BA Assignés',
        'target_mensuel': 'Objectif Mensuel',
        'total_new_add': 'GADD Total',
        'jours_actifs_15': 'Jours (>=15 BA Actifs)',
        'prime_fixe': 'Prime Fixe',
        'prime_variable': 'Prime Variable',
        'prime_15ba_jour': 'Prime 15 BA/Jour',
        'prime_mercenaire': 'Prime Mercenaire',
        'TOTAL COMMISSION': 'TOTAL À PAYER'
    }
    
    # On retire le msisdn_momo et region si non nécessaire visuellement, 
    # mais on garde l'essentiel pour la paie
    cols_to_keep = list(rename_cols.keys())
    df_synth_display = df_synth[cols_to_keep].rename(columns=rename_cols).copy()
    
    for r in dataframe_to_rows(df_synth_display, index=False, header=True):
        ws_synth.append(r)
        
    # Styles
    header_fill_1 = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    for cell in ws_synth[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill_1
        cell.alignment = Alignment(horizontal="center")
        
    for col in ws_synth.columns:
        ws_synth.column_dimensions[col[0].column_letter].width = 18
        
    # ========= ONGLET 2 : DÉTAILS BAs =========
    ws_detail = wb.create_sheet(title="Détails Agents")
    df_detail_display = df_detail.rename(columns={
        'superviseur': 'Superviseur',
        'agent': 'Nom Agent (BA)',
        'type_agent': 'Type Agent',
        'total_gadd': 'GADD Réalisé',
        'jours_actifs': 'Jours Actifs (>0 GADD)'
    })
    
    for r in dataframe_to_rows(df_detail_display, index=False, header=True):
        ws_detail.append(r)
        
    header_fill_2 = PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid")
    for cell in ws_detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill_2
        
    for col in ws_detail.columns:
        ws_detail.column_dimensions[col[0].column_letter].width = 22

    # Optionnel : Figer la première ligne
    ws_synth.freeze_panes = "A2"
    ws_detail.freeze_panes = "A2"
    
    wb.save(out_file)
    print(f"✓ Fichier généré : {out_file}")
    print("\nAperçu des Totaux calculés par l'algorithme SQL :")
    for _, row in df_synth_display.iterrows():
        print(f"  > {row['Superviseur']} ({row['Catégorie']}) --> {row['TOTAL À PAYER']:,.0f} FCFA")

if __name__ == "__main__":
    import sys
    # On permet de passer le mois en argument, sinon par défaut on prend février 2026
    target = sys.argv[1] if len(sys.argv) > 1 else "2026-02"
    generate_report(target)
