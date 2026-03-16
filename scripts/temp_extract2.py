import openpyxl
file_path = r"d:\LKA\Perf_commissions\Point Financier Superviseur BA Dec 2025.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
wb_v = openpyxl.load_workbook(file_path, data_only=True)

ws = wb['Variable 15 BA par Jour']
ws_v = wb_v['Variable 15 BA par Jour']
print("Variable 15 BA par Jour: AK2 Formula:", ws['AK2'].value, "Value:", ws_v['AK2'].value)
print("Variable 15 BA par Jour: Row 3/4 headers and formulas mapping:")
for col in range(5, ws.max_column + 1):
    h = ws_v.cell(3, col).value
    c = ws.cell(4, col)
    v = ws_v.cell(4, col)
    if c.value or v.value:
        print(f"Col {col} ({c.coordinate}) - Header '{str(h)[:15]}': {c.value} -> {v.value}")
