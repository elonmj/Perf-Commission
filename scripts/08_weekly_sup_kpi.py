import sys
import os
import json
import base64
import mimetypes
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
import numpy as np
from sqlalchemy import text
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from connections.config import MYSQL_DATABASE, TABLE_DAILY_GADD, RECIPIENTS_WEEKLY_SUP_KPI
from connections.connect import make_engine
import importlib

fetch_mail = importlib.import_module("scripts.00_fetch_mail")
get_gmail_service = fetch_mail.get_gmail_service

TRACKER_FILE = ROOT / "data" / "processed_sup_kpi_weeks.json"

def get_target_week(max_date, force_current=False):
    if force_current:
        days_since_sunday = (max_date.weekday() + 1) % 7
        start_date = max_date - timedelta(days=days_since_sunday)
        end_date = max_date
    else:
        days_since_saturday = (max_date.weekday() + 2) % 7
        last_saturday = max_date - timedelta(days=days_since_saturday)
        last_sunday = last_saturday - timedelta(days=6)
        start_date = last_sunday
        end_date = last_saturday
    return start_date, end_date

def load_processed_weeks():
    if not TRACKER_FILE.exists(): return []
    try: return json.loads(TRACKER_FILE.read_text(encoding="utf-8"))
    except: return []

def save_processed_week(week_id):
    TRACKER_FILE.parent.mkdir(exist_ok=True)
    weeks = load_processed_weeks()
    if week_id not in weeks:
        weeks.append(week_id)
        TRACKER_FILE.write_text(json.dumps(weeks), encoding="utf-8")

def send_email_with_attachment(service, subject, html_content, attachment_path, to_emails):
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = "me"
    msg["To"] = to_emails

    msg.add_alternative(html_content, subtype='html')

    path = Path(attachment_path)
    if path.exists():
        ctype, encoding = mimetypes.guess_type(str(path))
        if ctype is None or encoding is not None:
            ctype = 'application/octet-stream'
        maintype, subtype = ctype.split('/', 1)
        with open(path, 'rb') as f:
            msg.add_attachment(f.read(), maintype=maintype, subtype=subtype, filename=path.name)
            
    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    service.users().messages().send(userId="me", body={"raw": raw_message}).execute()

def list_available_months(engine):
    query = "SELECT DISTINCT perf_month FROM vw_commission_superviseur ORDER BY perf_month DESC"
    res = pd.read_sql(query, engine)
    return res['perf_month'].tolist()

def format_excel_headers(ws, fill_color="00B050", font_color="FFFFFF"):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color=font_color)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    ws.freeze_panes = "A2"

def generate_weekly_sup_kpi(engine, target_month, start_week, end_week):
    months = list_available_months(engine)
    if target_month not in months:
        if months:
            target_month = months[0]
        else:
            return None

    # Fetch without any prices (sans prix), inclure la région
    query_synth = f"SELECT superviseur_name, region, type_superviseur, perf_month, nb_ba_total, target_mensuel, total_new_add, jours_actifs_15 FROM vw_commission_superviseur WHERE perf_month = '{target_month}'"
    df_synth = pd.read_sql(query_synth, engine)
    if df_synth.empty: return None
    
    output_dir = os.path.join(ROOT, "outputs")
    os.makedirs(output_dir, exist_ok=True)
    out_file = os.path.join(output_dir, f"Weekly_Sup_KPIs_{start_week}_to_{end_week}.xlsx")
    
    wb = Workbook()
    ws_synth = wb.active
    ws_synth.title = "Progression Mensuelle KPIs"
    
    rename_cols = {
        'superviseur_name': 'Superviseur', 
        'region': 'Region',
        'type_superviseur': 'Catégorie', 
        'perf_month': 'Mois',
        'nb_ba_total': 'Nb BA Assignés', 
        'target_mensuel': 'Objectif GADD Mensuel', 
        'total_new_add': 'GADD Total MTD',
        'jours_actifs_15': 'Jours Actifs (>=15 BA)'
    }
    df_synth_display = df_synth.rename(columns=rename_cols).copy()
    
    # Calculate % progress with division by zero protection
    df_synth_display['% Objectif GADD'] = (df_synth_display['GADD Total MTD'] / df_synth_display['Objectif GADD Mensuel'].replace(0, np.nan)).fillna(0)
    
    # Calculer le score de performance par région
    region_scores = df_synth_display.groupby('Region')['% Objectif GADD'].mean().reset_index()
    region_scores.rename(columns={'% Objectif GADD': 'REGION_SCORE'}, inplace=True)
    
    # Fusionner pour trier (pire région d'abord, puis pire superviseur)
    df_synth_display = df_synth_display.merge(region_scores, on='Region')
    df_synth_display = df_synth_display.sort_values(by=['REGION_SCORE', 'Region', '% Objectif GADD'], ascending=[True, True, True])

    # Reorder columns
    cols = ['Region', 'Superviseur', 'Catégorie', 'Mois', 'Nb BA Assignés', 'Objectif GADD Mensuel', 'GADD Total MTD', '% Objectif GADD', 'Jours Actifs (>=15 BA)']
    if set(cols).issubset(df_synth_display.columns):
        df_synth_display = df_synth_display[cols]
    
    # Écriture manuelle pour insérer les lignes vides au changement de région
    for c_idx, col_name in enumerate(cols, 1):
        ws_synth.cell(row=1, column=c_idx, value=col_name)
        
    current_row = 2
    last_region = None
    
    for _, row_data in df_synth_display.iterrows():
        region = row_data['Region']
        if last_region is not None and region != last_region:
            current_row += 1  # Ligne vide
        
        for c_idx, col_name in enumerate(cols, 1):
            val = row_data[col_name]
            cell = ws_synth.cell(row=current_row, column=c_idx, value=val)
            if col_name == '% Objectif GADD':
                cell.number_format = '0.00%'
                
        last_region = region
        current_row += 1
        
    format_excel_headers(ws_synth, fill_color="00B050")
    
    wb.save(out_file)
    return out_file

def main():
    print("=== 08_weekly_sup_kpi.py ===")
    engine = make_engine(MYSQL_DATABASE)

    with engine.connect() as conn:
        row = conn.execute(text(f"SELECT MAX(perf_date) FROM {TABLE_DAILY_GADD}")).fetchone()
        
    if not row or not row[0]: return

    max_date = row[0]
    is_force = os.environ.get("FORCE_SEND_ALL") == "1"
    start_date, end_date = get_target_week(max_date, is_force)
    week_id = f"{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}"
    
    processed_weeks = load_processed_weeks()
    if week_id in processed_weeks and not is_force:
        print(f"  Semaine {week_id} déjà traitée (Sup KPI).")
        sys.exit(0)

    target_month = end_date.strftime('%Y-%m')
    kpi_file = generate_weekly_sup_kpi(engine, target_month, start_date, end_date)
    
    if not kpi_file: return

    service = get_gmail_service()
    html_kpi = f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #333;">
        <h2 style="color: #00B050;">Progression des KPIs Superviseurs</h2>
        <p><strong>Période évaluée :</strong> Semaine du {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}</p>
        <p>Bonjour,</p>
        <p>Afin de savoir "où vous en êtes" sans les informations de commission, veuillez trouver l'avancement de vos objectifs BA réels (Objectif Mensuel, GADD Total généré, etc.).</p>
        <p>Le fichier est trié pour afficher les meilleures performances en haut du tableau.</p>
        <br>
        <p style="font-size: 12px; color: #666;"><em>Service Reporting Automatisé</em></p>
      </body>
    </html>
    """
    subject_kpi = f"🎯 KPIs de progression Superviseur ({start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m')})"
    send_email_with_attachment(service, subject_kpi, html_kpi, kpi_file, RECIPIENTS_WEEKLY_SUP_KPI)

    save_processed_week(week_id)
    print("  Email Weekly Sup KPI envoyé !")

if __name__ == "__main__":
    main()