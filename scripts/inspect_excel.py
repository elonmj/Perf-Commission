"""
scripts/inspect_excel.py -- Audit du fichier de commissions Excel
Génère un rapport JSON et texte sur la structure du fichier (feuilles, colonnes, données, règles).
"""

import sys
import argparse
import json
from pathlib import Path
from openpyxl import load_workbook

def inspect_excel(filepath: Path, strict=False):
    if not filepath.exists():
        print(f"File not found: {filepath}")
        sys.exit(1)
        
    print(f"Inspecting: {filepath.name}...")
    wb = load_workbook(filename=filepath, read_only=False, data_only=False)
    
    report = {
        "file": filepath.name,
        "sheets": []
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "freeze_panes": str(ws.freeze_panes),
            "merged_cells": len(ws.merged_cells.ranges),
            "headers": [],
            "formulas": [],
            "tariff_grid": []
        }
        
        # Read a few headers
        header_row = None
        for row_idx in range(1, min(ws.max_row + 1, 15)):
            if ws.cell(row=row_idx, column=1).value == "USER NAME":
                header_row = row_idx
                break
        
        if header_row:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col_idx).value
                if val:
                    sheet_info["headers"].append(val)
                    
            # Sample formulas from last row
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=ws.max_row, column=col_idx).value
                if isinstance(val, str) and val.startswith("="):
                    sheet_info["formulas"].append({
                        "col": col_idx,
                        "formula": val
                    })
                    
        # Read Tariff Grid if present
        for row_idx in range(1, 15):
            val = ws.cell(row=row_idx, column=1).value
            if isinstance(val, str) and "RAPPEL DES TARIFS" in val:
                # Read tariff grid headers
                grid_headers = [ws.cell(row=row_idx+1, column=c).value for c in range(1, 6)]
                sheet_info["tariff_grid"] = [h for h in grid_headers if h]
                break

        report["sheets"].append(sheet_info)
        
    # Validation against rules
    errors = []
    
    # 1. Sheets naming
    expected_sheets = ["New Add (GADD)", "New UserData (ADS)"]
    for s in expected_sheets:
        if s not in wb.sheetnames and s.replace(" ", "") not in [sn.replace(" ", "") for sn in wb.sheetnames]:
            errors.append(f"Missing sheet: {s}")
            
    # 2. Tariff Grid must NOT contain Canal
    for s in report["sheets"]:
        if any("canal" in str(h).lower() for h in s["tariff_grid"]):
            errors.append(f"Tariff grid contains 'Canal' in sheet {s['name']}")
            
    # 3. Freeze Panes
    for s in report["sheets"]:
        if s["freeze_panes"] not in ["B9", "B8", "B10", "B11", "B12", "B13", "B7"]: # depending on the rows
             errors.append(f"Unexpected freeze panes {s['freeze_panes']} in sheet {s['name']}")
             
    # Output report
    print("\n--- AUDIT SUMMARY ---")
    for s in report["sheets"]:
        print(f"\nSheet: {s['name']}")
        print(f"  Dimensions: {s['max_row']} rows x {s['max_column']} cols")
        print(f"  Headers: {', '.join(s['headers'][:5])}...")
        print(f"  Freeze Panes: {s['freeze_panes']}")
        print(f"  Tariff Header: {s['tariff_grid']}")
        print(f"  Formula count (totals): {len(s['formulas'])}")
        
    if errors:
        print("\n❌ ERRORS DETECTED:")
        for e in errors:
            print(f"  - {e}")
        if strict:
            sys.exit(1)
    else:
        print("\n✅ FILE LOOKS GOOD")
        
    return report

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="Path to specific Excel file")
    parser.add_argument("--glob", help="Glob pattern to inspect multiple files")
    parser.add_argument("--strict", action="store_true", help="Fail with exit code 1 if errors found")
    args = parser.parse_args()
    
    if args.file:
        files = [Path(args.file)]
    elif args.glob:
        files = list(Path(".").glob(args.glob))
    else:
        # Latest 3 outputs by default
        outputs = list(Path("outputs").glob("Variables*.xlsx")) + list(Path("outputs").glob("commission_*.xlsx"))
        outputs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        files = outputs[:3]
        
    if not files:
        print("No files found to inspect.")
        sys.exit(1)
        
    json_reports = []
    
    for f in files:
        rep = inspect_excel(f, strict=args.strict)
        json_reports.append(rep)
        
    out_dir = Path("reports")
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "latest_inspection.json"
    
    with open(out_path, "w", encoding="utf-8") as outf:
        json.dump(json_reports, outf, indent=2, ensure_ascii=False)
        
    print(f"\nJSON report saved to {out_path}")

if __name__ == "__main__":
    main()
