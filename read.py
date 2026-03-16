import openpyxl
from openpyxl.utils import get_column_letter

FILE_PATH = r"d:\LKA\Perf_commissions\Point Financier Superviseur BA Dec 2025.xlsx"

print(f"=== Analyse de: {FILE_PATH} ===")
wb = openpyxl.load_workbook(FILE_PATH, data_only=False)

# 1. Lister les feuilles
print("\n1. FEUILLES DISPONIBLES :")
for i, sheet in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet}")

# 2. Focus sur la feuille "Recap"
target_sheet = next((s for s in wb.sheetnames if "recap" in s.lower()), wb.sheetnames[0])
print(f"\n2. FEUILLE CIBLE: '{target_sheet}'")
ws = wb[target_sheet]

print(f"  Dimensions : {ws.dimensions}")
print(f"  Max Ligne : {ws.max_row}, Max Colonne : {ws.max_column}")
print(f"  Freeze Panes : {ws.freeze_panes}")

# 3. En-têtes (Lignes 1 à 3 pour repérer les fusions)
print("\n3. STRUCTURE DES EN-TÊTES (Lignes 1 à 3) :")
for row in range(1, 4):
    row_data = []
    for col in range(1, min(15, ws.max_column + 1)):
        val = str(ws.cell(row, col).value)[:25] if ws.cell(row, col).value else ""
        row_data.append(val.ljust(25))
    print(f"  L{row} | {' | '.join(row_data)}")

# 4. Formules utilisées
print("\n4. ÉCHANTILLON DE FORMULES :")
formula_count = 0
for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row)):
    for cell in row:
        if cell.data_type == 'f':
            formula_count += 1
            if formula_count <= 20:
                print(f"  {cell.coordinate}: {cell.value}")

print("\n(Terminé)")